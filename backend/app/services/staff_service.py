"""Staff service layer with business logic and helper functions."""

import csv
import os
import re
import unicodedata
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from uuid import uuid4

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func, or_, and_
from sqlalchemy.orm import Session

from app.core.security import get_password_hash, verify_password
from app.models.user import User
from app.models.product import Product
from app.models.order import Order
from app.models.order_item import OrderItem
from app.models.category import Category
from app.models.inventory_lot import InventoryLot
from app.models.store import Store
from app.models.donation_offer import DonationOffer
from app.models.donation_request import DonationRequest
from app.models.donation_request_item import DonationRequestItem
from app.models.delivery import Delivery
from app.models.delivery_partner import DeliveryPartner
from app.models.charity_organization import CharityOrganization


# ========== Helper Functions ==========
def _dict_row(row) -> dict:
    return dict(row._mapping)


def _status_label(expiry_date: date) -> str:
    today = date.today()
    if expiry_date < today:
        return "Het Han"
    if (expiry_date - today).days <= 7:
        return "Sap Het Han"
    return "Moi"


def _parse_date_input(value) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            raise ValueError("Ngay het han trong")
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(stripped, fmt).date()
            except ValueError:
                continue
    raise ValueError("Ngay het han khong hop le")


def _normalize_header(header: object) -> str:
    if header is None:
        return ""
    value = unicodedata.normalize("NFKD", str(header).strip().lower())
    value = "".join(char for char in value if not unicodedata.combining(char))
    value = value.replace(" ", "_")
    value = value.replace("-", "_")
    return value


def _normalize_status_value(raw_status: object, expiry_date: date) -> str:
    if raw_status is None:
        return _status_label(expiry_date)

    normalized = unicodedata.normalize("NFKD", str(raw_status).strip().lower())
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = normalized.replace("_", " ")

    status_map = {
        "moi": "Moi",
        "sap het han": "Sap Het Han",
        "het han": "Het Han",
    }
    return status_map.get(normalized, _status_label(expiry_date))


def _parse_non_negative_float(value: object, field_name: str) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None

    normalized = str(value).strip().replace(" ", "")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(",", "")
    elif "," in normalized and "." not in normalized:
        normalized = normalized.replace(",", ".")
    try:
        result = float(normalized)
    except ValueError as exc:
        raise ValueError(f"{field_name} khong hop le") from exc

    if result < 0:
        raise ValueError(f"{field_name} phai >= 0")
    return result


def _pick_field(row: dict[str, object], *keys: str) -> object:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def _build_product_sku(name: str) -> str:
    base = re.sub(r"[^a-z0-9]+", "", name.strip().lower())
    if not base:
        base = "product"
    return f"sku_{base[:20]}"


def _generate_unique_sku(db: Session, supermarket_id: int, product_name: str) -> str:
    base = _build_product_sku(product_name)
    candidate = base
    index = 1

    while True:
        exists = db.query(Product.id).filter(
            Product.supermarket_id == supermarket_id,
            Product.sku == candidate
        ).first()

        if not exists:
            return candidate

        index += 1
        candidate = f"{base}_{index}"


def _get_staff_scope(db: Session, user_id: int) -> dict[str, int]:
    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    role = (user.role or "").lower()
    if role != "store_staff":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Tai khoan khong co quyen staff.",
        )

    if not user.store_id or not user.supermarket_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tai khoan staff chua duoc gan store.",
        )

    return {"store_id": int(user.store_id), "supermarket_id": int(user.supermarket_id)}


def _resolve_or_create_category(db: Session, category_name: object) -> int | None:
    name = str(category_name or "").strip()
    if not name:
        return None

    existing = db.query(Category.id).filter(
        func.lower(Category.name) == func.lower(name)
    ).first()

    if existing:
        return int(existing.id)

    new_category = Category(name=name)
    db.add(new_category)
    db.flush()
    return int(new_category.id)


def _ensure_unique_sku_for_product(
    db: Session,
    supermarket_id: int,
    sku: str,
    current_product_id: int | None = None,
) -> None:
    query = db.query(Product.id).filter(
        Product.supermarket_id == supermarket_id,
        Product.sku == sku
    )
    
    if current_product_id is not None:
        query = query.filter(Product.id != current_product_id)
    
    conflict = query.first()

    if conflict:
        raise ValueError(f"SKU '{sku}' da ton tai")


def _resolve_or_create_product(db: Session, supermarket_id: int, product_name: str) -> int:
    name = product_name.strip()
    product = db.query(Product.id).filter(
        Product.supermarket_id == supermarket_id,
        func.lower(Product.name) == func.lower(name)
    ).first()

    if product:
        return int(product.id)

    sku = _generate_unique_sku(db, supermarket_id, name)
    new_product = Product(
        supermarket_id=supermarket_id,
        category_id=None,
        sku=sku,
        name=name,
        base_price=0,
        image_url=None
    )
    db.add(new_product)
    db.flush()
    return int(new_product.id)


def _upsert_product_from_import(
    db: Session,
    *,
    supermarket_id: int,
    product_name_raw: object,
    sku_raw: object = None,
    base_price_raw: object = None,
    category_name_raw: object = None,
) -> tuple[int, str]:
    product_name = str(product_name_raw or "").strip()
    if not product_name:
        raise ValueError("Ten san pham khong duoc trong")

    sku = str(sku_raw or "").strip()
    if not sku:
        sku = None

    base_price = _parse_non_negative_float(base_price_raw, "Don gia")
    category_id = _resolve_or_create_category(db, category_name_raw)

    product = None
    if sku:
        product = db.query(Product.id, Product.sku, Product.base_price, Product.category_id).filter(
            Product.supermarket_id == supermarket_id,
            Product.sku == sku
        ).first()

    if not product:
        product = db.query(Product.id, Product.sku, Product.base_price, Product.category_id).filter(
            Product.supermarket_id == supermarket_id,
            func.lower(Product.name) == func.lower(product_name)
        ).first()

    if product:
        current_id = int(product.id)
        next_sku = str(product.sku)
        if sku and sku != next_sku:
            _ensure_unique_sku_for_product(db, supermarket_id, sku, current_product_id=current_id)
            next_sku = sku

        next_base_price = float(base_price) if base_price is not None else float(product.base_price or 0)
        next_category_id = category_id if category_id is not None else product.category_id

        db.query(Product).filter(Product.id == current_id).update(
            {
                Product.sku: next_sku,
                Product.name: product_name,
                Product.base_price: next_base_price,
                Product.category_id: next_category_id,
            },
            synchronize_session=False
        )
        db.flush()
        return current_id, "updated"

    next_sku = sku or _generate_unique_sku(db, supermarket_id, product_name)
    _ensure_unique_sku_for_product(db, supermarket_id, next_sku)
    new_product = Product(
        supermarket_id=supermarket_id,
        category_id=category_id,
        sku=next_sku,
        name=product_name,
        base_price=float(base_price) if base_price is not None else 0,
        image_url=None
    )
    db.add(new_product)
    db.flush()
    return int(new_product.id), "created"


def _upsert_inventory_lot(
    db: Session,
    *,
    store_id: int,
    supermarket_id: int,
    lot_code: str,
    product_name: str,
    quantity: int,
    expiry_date: date,
    manual_status: object = None,
    product_id: int | None = None,
    manufacturing_date: date | None = None,
) -> str:
    # 1. Tim lo hang hien tai theo store va lot_code
    existing_lot = db.query(InventoryLot).filter(
        InventoryLot.store_id == store_id,
        InventoryLot.lot_code == lot_code
    ).first()

    if product_id is None:
        product_id = _resolve_or_create_product(db, supermarket_id, product_name)
    
    lot_status = _normalize_status_value(manual_status, expiry_date)

    if existing_lot:
        # 2. KIỂM TRA CHẶN GHI ĐÈ SẢN PHẨM: Nếu trùng mã lô nhưng khác sản phẩm -> Báo lỗi
        if existing_lot.product_id != product_id:
            old_p = db.query(Product.name).filter(Product.id == existing_lot.product_id).first()
            new_p = db.query(Product.name).filter(Product.id == product_id).first()
            old_name = old_p.name if old_p else f"ID:{existing_lot.product_id}"
            new_name = new_p.name if new_p else f"ID:{product_id}"
            raise ValueError(
                f"Mã lô '{lot_code}' đã tồn tại cho sản phẩm '{old_name}'. "
                f"Không thể dùng cho sản phẩm khác là '{new_name}'. "
                "Vui lòng kiểm tra lại hoặc sử dụng mã lô khác."
            )

        # 3. Kiểm tra nếu trùng mã lô, ngày hết hạn và ngày sản xuất có khớp không
        if existing_lot.expiry_date != expiry_date or existing_lot.manufacturing_date != manufacturing_date:
            raise ValueError(
                f"Mã lô '{lot_code}' đã tồn tại với thông tin ngày khác "
                f"(NSX: {existing_lot.manufacturing_date}, HSD: {existing_lot.expiry_date}). "
                f"Không thể cộng dồn với dữ liệu mới (NSX: {manufacturing_date}, HSD: {expiry_date}). "
                "Vui lòng kiểm tra lại hoặc dùng mã lô khác."
            )
        
        # 4. Nếu khớp tất cả thông tin -> Thực hiện CỘNG DỒN số lượng
        new_qty = (existing_lot.qty_on_hand or 0) + quantity
        new_imported = (existing_lot.qty_imported or existing_lot.qty_on_hand or 0) + quantity
        
        db.query(InventoryLot).filter(InventoryLot.id == existing_lot.id).update(
            {
                InventoryLot.qty_on_hand: new_qty,
                InventoryLot.qty_imported: new_imported,
                InventoryLot.status: lot_status,
            },
            synchronize_session=False
        )
        db.flush()
        return "updated"

    # 4. Nếu chưa có -> Tạo mới
    new_lot = InventoryLot(
        store_id=store_id,
        product_id=product_id,
        lot_code=lot_code,
        expiry_date=expiry_date,
        manufacturing_date=manufacturing_date,
        qty_on_hand=quantity,
        qty_imported=quantity,
        qty_reserved=0,
        status=lot_status
    )
    db.add(new_lot)
    db.flush()
    return "created"


def _extract_rows_from_xlsx(file_bytes: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_header(h) for h in rows[0]]
    data_rows: list[dict[str, object]] = []
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item: dict[str, object] = {}
        for idx, value in enumerate(row):
            if idx < len(headers) and headers[idx]:
                item[headers[idx]] = value
        data_rows.append(item)
    return data_rows


def _extract_rows_from_csv(file_bytes: bytes) -> list[dict[str, object]]:
    text_stream = StringIO(file_bytes.decode("utf-8-sig"))
    reader = csv.DictReader(text_stream)
    rows = []
    for row in reader:
        normalized = {_normalize_header(k): v for k, v in row.items()}
        if any(str(v or "").strip() for v in normalized.values()):
            rows.append(normalized)
    return rows


def _read_import_rows(upload: UploadFile, file_bytes: bytes) -> list[dict[str, object]]:
    filename = (upload.filename or "").lower()
    if filename.endswith(".xlsx"):
        return _extract_rows_from_xlsx(file_bytes)
    if filename.endswith(".csv"):
        return _extract_rows_from_csv(file_bytes)

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Chi ho tro file .xlsx hoac .csv",
    )


# ========== Profile Business Logic ==========
def get_staff_profile(db: Session, user_id: int) -> dict:
    user = db.query(
        User.id,
        User.username,
        User.email,
        User.full_name,
        User.phone,
        User.role,
        User.store_id,
        User.supermarket_id,
        Store.name.label('store_name'),
        Store.location.label('store_address')
    ).outerjoin(
        Store, Store.id == User.store_id
    ).filter(
        User.id == user_id
    ).first()

    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    return {
        "email": user.email,
        "fullName": user.full_name,
        "phone": user.phone,
        "role": user.role,
        "storeName": user.store_name or "Chưa gán cửa hàng",
        "storeAddress": user.store_address or "",
    }


def update_staff_profile(db: Session, user_id: int, full_name: str, email: str, phone: str) -> dict:
    full_name = full_name.strip()
    email = email.strip().lower()
    phone = phone.strip()

    if not full_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Họ tên không được để trống")
    if not email:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email không được để trống")

    existing_email = db.query(User.id).filter(
        User.email == email,
        User.id != user_id
    ).first()
    if existing_email:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email đã được sử dụng")

    if phone:
        existing_phone = db.query(User.id).filter(
            User.phone == phone,
            User.id != user_id
        ).first()
        if existing_phone:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Số điện thoại đã được sử dụng")

    db.query(User).filter(User.id == user_id).update(
        {
            User.full_name: full_name,
            User.email: email,
            User.phone: phone
        },
        synchronize_session=False
    )
    db.commit()
    
    # Return the updated profile
    return get_staff_profile(db, user_id)


def change_staff_password(db: Session, user_id: int, current_password: str, new_password: str) -> dict:
    current_password = current_password or ""
    new_password = new_password or ""

    if not current_password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Vui lòng nhập mật khẩu hiện tại"
        )

    if len(new_password) < 6:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Mật khẩu mới phải có ít nhất 6 ký tự"
        )

    row = db.query(User.password_hash).filter(User.id == user_id).first()

    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Không tìm thấy người dùng"
        )

    if not verify_password(current_password, row.password_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Mật khẩu hiện tại không đúng"
        )

    db.query(User).filter(User.id == user_id).update(
        {User.password_hash: get_password_hash(new_password)},
        synchronize_session=False
    )
    db.commit()

    return {"message": "Đổi mật khẩu thành công!", "success": True}


# ========== Orders Business Logic ==========
def list_staff_orders(db: Session, store_id: int) -> dict:
    rows = db.query(
        Order.id, Order.status, Order.total_amount, Order.payment_status, Order.created_at, Order.delivered_at,
        User.full_name.label('customer_name')
    ).join(
        User, User.id == Order.customer_id
    ).filter(
        Order.store_id == store_id,
        Order.status != 'cancelled',
        or_(
            Order.payment_status == 'paid',
            Order.payment_method == 'cod'
        )
    ).order_by(
        Order.created_at.desc()
    ).limit(100).all()

    items = []
    for row in rows:
        items.append({
            "id": f"DH-{row.id}",
            "orderId": row.id,
            "customer": row.customer_name,
            "status": row.status,
            "amount": f"{float(row.total_amount or 0):,.0f} VNĐ" if row.total_amount else "0 VNĐ",
            "paymentStatus": row.payment_status,
            "createdAt": row.created_at.strftime("%d/%m/%Y %H:%M"),
            "deliveredAt": row.delivered_at.strftime("%d/%m/%Y %H:%M") if row.delivered_at else None,
        })

    return {"items": items}


# ========== Delivery Assignment Functions ==========
def _find_available_delivery_partner(db: Session, store_id: int):
    partner = db.query(
        DeliveryPartner.id,
        DeliveryPartner.user_id,
        func.count(Delivery.id).label("active_count")
    ).outerjoin(
        Delivery,
        and_(
            Delivery.delivery_partner_id == DeliveryPartner.id,
            Delivery.status.in_(["assigned", "picking_up", "delivering"])
        )
    ).join(
        User,
        User.id == DeliveryPartner.user_id
    ).filter(
        User.is_active == True
    ).group_by(
        DeliveryPartner.id
    ).order_by(
        func.count(Delivery.id).asc()
    ).first()
    
    return partner


def _generate_delivery_code(db: Session) -> str:
    timestamp = datetime.now().strftime("%Y%m%d")
    count = db.query(func.count(Delivery.id)).filter(
        Delivery.delivery_code.like(f"GH-{timestamp}-%")
    ).scalar() or 0
    return f"GH-{timestamp}-{str(count + 1).zfill(4)}"


def _create_delivery_for_order(db: Session, order_id: int, store_id: int) -> dict:
    # Get order and customer
    order_row = db.query(
        Order.id,
        Order.customer_id,
        Order.shipping_address,
        User.full_name.label("customer_name"),
        User.phone.label("customer_phone"),
        Store.location.label("store_location")
    ).join(
        User, User.id == Order.customer_id
    ).join(
        Store, Store.id == Order.store_id
    ).filter(
        Order.id == order_id,
        Order.store_id == store_id
    ).first()
    
    if not order_row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Đơn hàng không tồn tại"
        )
    
    # Check if delivery already exists for this order
    existing_delivery = db.query(Delivery).filter(
        Delivery.order_id == order_id
    ).first()
    
    if existing_delivery:
        return {"delivery_id": existing_delivery.id}
    
    # Find available delivery partner
    partner = _find_available_delivery_partner(db, store_id)
    
    if not partner:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Không có nhân viên giao hàng khả dụng"
        )
    
    # Create delivery
    delivery_code = _generate_delivery_code(db)
    
    new_delivery = Delivery(
        delivery_code=delivery_code,
        order_id=order_id,
        store_id=store_id,
        delivery_partner_id=partner.id,
        receiver_name=order_row.customer_name,
        receiver_phone=order_row.customer_phone,
        receiver_address=order_row.shipping_address or order_row.store_location,
        status="assigned"
    )
    
    db.add(new_delivery)
    db.flush()
    delivery_id = new_delivery.id

    db.commit()

    return {"delivery_id": delivery_id, "delivery_code": delivery_code}


def _create_delivery_for_donation_request(db: Session, donation_request_id: int, store_id: int) -> dict:
    # Get donation request with related details (via DonationRequestItem)
    # Include charity's address for delivery destination
    request_row = db.query(
        DonationRequest.id,
        DonationRequest.charity_id,
        User.full_name.label("charity_name"),
        User.phone.label("charity_phone"),
        CharityOrganization.address.label("charity_address")
    ).join(
        User, User.id == DonationRequest.charity_id
    ).join(
        CharityOrganization, CharityOrganization.user_id == DonationRequest.charity_id
    ).filter(
        DonationRequest.id == donation_request_id
    ).first()
    
    if not request_row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Donation request không tồn tại"
        )
    
    # Check if delivery already exists for this donation request from this store
    existing_delivery = db.query(Delivery).filter(
        Delivery.donation_request_id == donation_request_id,
        Delivery.store_id == store_id
    ).first()
    
    if existing_delivery:
        return {"delivery_id": existing_delivery.id, "delivery_code": existing_delivery.delivery_code}
    
    # Find available delivery partner
    partner = _find_available_delivery_partner(db, store_id)
    
    if not partner:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Không có nhân viên giao hàng khả dụng"
        )
    
    # Create delivery with donation_request_id
    delivery_code = _generate_delivery_code(db)
    
    new_delivery = Delivery(
        delivery_code=delivery_code,
        donation_request_id=donation_request_id,
        store_id=store_id,
        delivery_partner_id=partner.id,
        receiver_name=request_row.charity_name,
        receiver_phone=request_row.charity_phone,
        receiver_address=request_row.charity_address or "Chua co dia chi",
        status="assigned"
    )
    
    db.add(new_delivery)
    db.flush()
    delivery_id = new_delivery.id

    db.commit()

    return {"delivery_id": delivery_id, "delivery_code": delivery_code}



def update_staff_order_status(db: Session, order_id: int, store_id: int, new_status: str, user_id: int = None) -> dict:
    new_status = (new_status or "").strip().lower()

    # Staff chỉ được update thành "ready" - đúng nghiệp vụ
    if new_status != "ready":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Staff chỉ được cập nhật đơn hàng thành 'ready'. Trạng thái khác do Delivery Partner xác nhận."
        )

    # Check if order exists in staff's store
    order = db.query(Order.id, Order.status).filter(
        Order.id == order_id,
        Order.store_id == store_id
    ).first()

    if not order:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Đơn hàng không tồn tại")

    # Only allow update from pending or preparing state to ready
    if order.status not in ["pending", "preparing"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Không thể cập nhật từ trạng thái '{order.status}' thành 'ready'"
        )

    # Deduct stock as the order is marked 'ready' (packed and handed to delivery)
    order_items = db.query(OrderItem).filter(OrderItem.order_id == order_id).all()
    from app.services.audit_service import log_action
    
    for item in order_items:
        if item.lot_id:
            lot = db.query(InventoryLot).filter(InventoryLot.id == item.lot_id).with_for_update().first()
            if lot:
                # KIỂM TRA SAI LỆCH KHO (Race Condition Check)
                on_hand = lot.qty_on_hand or 0
                if on_hand < item.quantity:
                    # Ghi log cảnh báo lệch kho nghiêm trọng để Admin kiểm tra sau
                    log_action(
                        db, user_id=user_id, store_id=store_id,
                        action="INVENTORY_MISMATCH",
                        entity_type="order_item",
                        entity_id=item.id,
                        old_value={"qty_on_hand": on_hand},
                        new_value={"required": item.quantity, "msg": f"Xuất kho vượt quá tồn thực tế cho đơn hàng DH-{order_id}"}
                    )

                # Trừ tồn kho thực tế và giải phóng lượng giữ chỗ
                lot.qty_on_hand = max(0, on_hand - item.quantity)
                lot.qty_reserved = max(0, (lot.qty_reserved or 0) - item.quantity)

    # Cập nhật trạng thái đơn hàng
    update_result = db.query(Order).filter(
        Order.id == order_id,
        Order.store_id == store_id
    ).update({Order.status: "ready"}, synchronize_session=False)
    
    if update_result == 0:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Đơn hàng không tồn tại hoặc không thuộc cửa hàng")
    
    db.commit()

    # Audit log for order approval
    if user_id:
        from app.services.audit_service import log_action
        from app.core.audit_actions import APPROVE_ORDER, ENTITY_ORDER
        log_action(db, user_id=user_id, store_id=store_id,
                   action=APPROVE_ORDER, entity_type=ENTITY_ORDER, entity_id=order_id,
                   old_value={"status": order.status}, new_value={"status": "ready"})

    response = {"success": True, "status": "ready"}

    # Auto-create delivery when order is ready
    try:
        delivery_info = _create_delivery_for_order(db, order_id, store_id)
        response["delivery"] = {
            "delivery_id": delivery_info.get("delivery_id"),
            "delivery_code": delivery_info.get("delivery_code"),
            "message": "Đơn giao hàng đã được tạo và gửi cho nhân viên giao hàng"
        }
    except HTTPException as e:
        # If delivery creation fails, log it but don't fail the order status update
        response["delivery_error"] = e.detail

    return response


def get_staff_order_detail(db: Session, order_id: int, store_id: int) -> dict:
    # Lấy thông tin đơn hàng + khách hàng
    order_row = db.query(
        Order.id, Order.status, Order.total_amount, Order.payment_method, Order.payment_status,
        Order.created_at, Order.delivered_at,
        User.full_name.label("customer_name"),
        User.phone.label("customer_phone")
    ).join(User, User.id == Order.customer_id).filter(
        Order.id == order_id,
        Order.store_id == store_id
    ).first()

    if not order_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Đơn hàng không tồn tại")

    # Lấy danh sách sản phẩm trong đơn hàng
    item_rows = db.query(
        OrderItem.quantity, OrderItem.unit_price,
        Product.name.label("product_name")
    ).join(Product, Product.id == OrderItem.product_id).filter(
        OrderItem.order_id == order_id
    ).all()

    items = []
    for row in item_rows:
        items.append({
            "productName": row.product_name,
            "quantity": row.quantity,
            "unitPrice": f"{float(row.unit_price):,.0f} VNĐ",
            "subtotal": f"{float(row.unit_price * row.quantity):,.0f} VNĐ",
        })

    return {
        "id": f"DH-{order_row.id}",
        "orderId": order_row.id,
        "customer": order_row.customer_name,
        "phone": order_row.customer_phone or "—",
        "status": order_row.status,
        "amount": f"{float(order_row.total_amount or 0):,.0f} VNĐ" if order_row.total_amount else "0 VNĐ",
        "paymentMethod": order_row.payment_method or "—",
        "paymentMethodText": "Tiền mặt" if order_row.payment_method == "cod"
                             else ("VNPay" if order_row.payment_method == "vnpay"
                             else ("Ví SEIMS" if order_row.payment_method == "wallet" else "—")),
        "paymentStatus": order_row.payment_status,
        "createdAt": order_row.created_at.strftime("%d/%m/%Y %H:%M"),
        "deliveredAt": order_row.delivered_at.strftime("%d/%m/%Y %H:%M") if order_row.delivered_at else None,
        "items": items,
    }


# ========== Categories Business Logic ==========
def staff_category_stats(db: Session, store_id: int) -> dict:
    rows = db.query(
        func.coalesce(Category.name, 'Khác').label('category_name'),
        func.count(InventoryLot.id).label('lot_count')
    ).join(
        Product, Product.id == InventoryLot.product_id
    ).outerjoin(
        Category, Category.id == Product.category_id
    ).filter(
        InventoryLot.store_id == store_id,
        InventoryLot.expiry_date >= date.today(),
        InventoryLot.expiry_date <= date.today() + timedelta(days=7)
    ).group_by(
        Category.id, Category.name
    ).order_by(
        func.count(InventoryLot.id).desc()
    ).limit(10).all()

    total = sum(row.lot_count for row in rows) or 1
    items = []
    for row in rows:
        items.append({
            "name": row.category_name,
            "percent": int((row.lot_count / total) * 100),
        })

    return {"items": items}


def list_categories(db: Session, supermarket_id: int) -> dict:
    rows = db.query(
        Category.id,
        Category.name,
        func.count(Product.id).label('product_count')
    ).outerjoin(
        Product, and_(Product.category_id == Category.id, Product.supermarket_id == supermarket_id)
    ).group_by(
        Category.id, Category.name
    ).order_by(
        Category.name.asc()
    ).all()

    items = [
        {"id": row.id, "name": row.name, "productCount": row.product_count}
        for row in rows
    ]

    return {"items": items}


def create_category(db: Session, name: str, user_id: int = None, store_id: int = None) -> dict:
    name = name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên danh mục không được trống")

    existing = db.query(Category.id).filter(
        func.lower(Category.name) == func.lower(name)
    ).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Danh mục đã tồn tại")

    new_category = Category(name=name)
    db.add(new_category)
    db.flush()
    category_id = new_category.id
    db.commit()

    if user_id and store_id:
        from app.services.audit_service import log_action
        from app.core.audit_actions import CREATE_CATEGORY, ENTITY_CATEGORY
        log_action(db, user_id=user_id, store_id=store_id,
                   action=CREATE_CATEGORY, entity_type=ENTITY_CATEGORY, entity_id=category_id,
                   new_value={"name": name})

    return {"message": "Tạo danh mục thành công", "name": name}


def update_category(db: Session, category_id: int, name: str, user_id: int = None, store_id: int = None) -> dict:
    name = name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên danh mục không được trống")

    # Get old value
    old_category = db.query(Category).filter(Category.id == category_id).first()
    old_name = old_category.name if old_category else None

    existing = db.query(Category.id).filter(
        func.lower(Category.name) == func.lower(name),
        Category.id != category_id
    ).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Danh mục đã tồn tại")

    db.query(Category).filter(Category.id == category_id).update(
        {Category.name: name},
        synchronize_session=False
    )
    db.commit()

    if user_id and store_id:
        from app.services.audit_service import log_action
        from app.core.audit_actions import UPDATE_CATEGORY, ENTITY_CATEGORY
        log_action(db, user_id=user_id, store_id=store_id,
                   action=UPDATE_CATEGORY, entity_type=ENTITY_CATEGORY, entity_id=category_id,
                   old_value={"name": old_name}, new_value={"name": name})

    return {"message": "Cập nhật danh mục thành công"}


def delete_category(db: Session, category_id: int, user_id: int = None, store_id: int = None) -> dict:
    old_category = db.query(Category).filter(Category.id == category_id).first()
    old_name = old_category.name if old_category else None

    has_products = db.query(func.count(Product.id)).filter(
        Product.category_id == category_id
    ).scalar() or 0

    if has_products > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Không thể xóa: có {has_products} sản phẩm đang sử dụng danh mục này"
        )

    db.query(Category).filter(Category.id == category_id).delete()
    db.commit()

    if user_id and store_id:
        from app.services.audit_service import log_action
        from app.core.audit_actions import DELETE_CATEGORY, ENTITY_CATEGORY
        log_action(db, user_id=user_id, store_id=store_id,
                   action=DELETE_CATEGORY, entity_type=ENTITY_CATEGORY, entity_id=category_id,
                   old_value={"name": old_name})

    return {"message": "Xóa danh mục thành công"}


# ========== Products Business Logic ==========
def list_products(db: Session, supermarket_id: int, category_filter: int | None, search: str | None) -> dict:
    base_query = db.query(
        Product.id, Product.sku, Product.name, Product.base_price, Product.image_url,
        Category.name.label("category_name"), Category.id.label("category_id"),
        func.coalesce(func.sum(InventoryLot.qty_on_hand), 0).label("total_stock")
    ).outerjoin(Category, Category.id == Product.category_id)\
     .outerjoin(InventoryLot, InventoryLot.product_id == Product.id)\
     .filter(Product.supermarket_id == supermarket_id)

    if category_filter is not None:
        base_query = base_query.filter(Product.category_id == category_filter)

    if search:
        base_query = base_query.filter(
            or_(Product.name.ilike(f"%{search}%"), Product.sku.ilike(f"%{search}%"))
        )

    rows = base_query.group_by(
        Product.id, Product.sku, Product.name, Product.base_price, Product.image_url, 
        Category.name, Category.id
    ).order_by(Product.name.asc()).all()

    items = [
        {
            "id": row.id,
            "sku": row.sku,
            "name": row.name,
            "basePrice": float(row.base_price),
            "imageUrl": row.image_url,
            "categoryName": row.category_name or "Chưa phân loại",
            "categoryId": row.category_id,
            "totalStock": int(row.total_stock),
        }
        for row in rows
    ]

    return {"items": items}


def create_product(db: Session, supermarket_id: int, store_id: int, user_id: int,
                  name: str, sku: str, base_price: float,
                  category_id: int | None, image_url: str | None) -> dict:
    name = name.strip()
    sku = sku.strip()
    image_url = (image_url or "").strip() or None

    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên sản phẩm không được trống")
    if not sku:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SKU không được trống")
    if base_price is None or float(base_price) < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Giá không hợp lệ")

    existing = db.query(Product.id).filter(
        Product.supermarket_id == supermarket_id,
        Product.sku == sku
    ).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SKU đã tồn tại")

    new_product = Product(
        supermarket_id=supermarket_id,
        category_id=category_id if category_id else None,
        sku=sku,
        name=name,
        base_price=float(base_price),
        image_url=image_url
    )
    db.add(new_product)
    db.flush()
    product_id = new_product.id
    db.commit()

    from app.services.audit_service import log_action
    from app.core.audit_actions import CREATE_PRODUCT, ENTITY_PRODUCT
    log_action(db, user_id=user_id, store_id=store_id,
               action=CREATE_PRODUCT, entity_type=ENTITY_PRODUCT, entity_id=product_id,
               new_value={
                   "name": name,
                   "sku": sku,
                   "base_price": float(base_price),
                   "category_id": category_id
               })

    return {"message": "Tạo sản phẩm thành công"}


def update_product(db: Session, product_id: int, supermarket_id: int,
                  store_id: int, user_id: int,
                  name: str, base_price: float,
                  category_id: int | None, image_url: str | None) -> dict:
    name = name.strip()
    image_url = (image_url or "").strip() or None

    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên sản phẩm không được trống")
    if base_price is None or float(base_price) < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Giá không hợp lệ")

    product = db.query(Product).filter(
        Product.id == product_id,
        Product.supermarket_id == supermarket_id
    ).first()
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sản phẩm không tìm thấy")

    old_price = float(product.base_price)
    new_price = float(base_price)
    old_value = {
        "name": product.name,
        "base_price": old_price,
        "category_id": product.category_id
    }

    db.query(Product).filter(Product.id == product_id).update({
        Product.name: name,
        Product.base_price: new_price,
        Product.category_id: category_id if category_id else None,
        Product.image_url: image_url
    }, synchronize_session=False)
    db.commit()

    new_value = {
        "name": name,
        "base_price": new_price,
        "category_id": category_id
    }

    from app.services.audit_service import log_action
    from app.core.audit_actions import UPDATE_PRODUCT, UPDATE_PRICE, ENTITY_PRODUCT
    log_action(db, user_id=user_id, store_id=store_id,
               action=UPDATE_PRODUCT, entity_type=ENTITY_PRODUCT, entity_id=product_id,
               old_value=old_value, new_value=new_value)

    if old_price != new_price:
        log_action(db, user_id=user_id, store_id=store_id,
                   action=UPDATE_PRICE, entity_type=ENTITY_PRODUCT, entity_id=product_id,
                   old_value={"base_price": old_price}, new_value={"base_price": new_price})

    return {"message": "Cập nhật sản phẩm thành công"}


def delete_product(db: Session, product_id: int, supermarket_id: int,
                  store_id: int, user_id: int) -> dict:
    product = db.query(Product).filter(
        Product.id == product_id,
        Product.supermarket_id == supermarket_id
    ).first()
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sản phẩm không tìm thấy")

    # Save old value before delete
    old_value = {
        "name": product.name,
        "sku": product.sku,
        "base_price": float(product.base_price)
    }

    has_lots = db.query(func.count(InventoryLot.id)).filter(
        InventoryLot.product_id == product_id
    ).scalar() or 0

    if has_lots > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Không thể xóa: có {has_lots} lô hàng liên quan đến sản phẩm này"
        )

    db.query(Product).filter(Product.id == product_id).delete()
    db.commit()

    from app.services.audit_service import log_action
    from app.core.audit_actions import DELETE_PRODUCT, ENTITY_PRODUCT
    log_action(db, user_id=user_id, store_id=store_id,
               action=DELETE_PRODUCT, entity_type=ENTITY_PRODUCT, entity_id=product_id,
               old_value=old_value)

    return {"message": "Xóa sản phẩm thành công"}


def list_product_categories(db: Session, supermarket_id: int) -> dict:
    rows = db.query(Category.id, Category.name).distinct().join(
        Product, Product.category_id == Category.id
    ).filter(
        Product.supermarket_id == supermarket_id
    ).order_by(Category.name.asc()).all()

    items = [{"id": row.id, "name": row.name} for row in rows]
    return {"items": items}


# ========== Dashboard Business Logic ==========
def staff_dashboard_summary(db: Session, store_id: int) -> dict:
    total_lots = db.query(func.count(InventoryLot.id)).filter(
        InventoryLot.store_id == store_id,
        InventoryLot.qty_on_hand > 0
    ).scalar() or 0

    total_inventory_qty = db.query(func.sum(InventoryLot.qty_on_hand)).filter(
        InventoryLot.store_id == store_id
    ).scalar() or 0

    near_expiry = db.query(func.count(InventoryLot.id)).filter(
        InventoryLot.store_id == store_id,
        InventoryLot.expiry_date >= date.today(),
        InventoryLot.expiry_date <= date.today() + timedelta(days=7),
        InventoryLot.qty_on_hand > 0
    ).scalar() or 0

    low_stock = db.query(func.count(InventoryLot.id)).filter(
        InventoryLot.store_id == store_id,
        InventoryLot.qty_on_hand > 0,
        InventoryLot.qty_on_hand <= 5
    ).scalar() or 0

    orders_today = db.query(func.count(Order.id)).filter(
        Order.store_id == store_id,
        Order.status != 'cancelled',
        or_(
            Order.payment_status == 'paid',
            Order.payment_method == 'cod'
        ),
        func.date(Order.created_at) == date.today()
    ).scalar() or 0

    orders_pending = db.query(func.count(Order.id)).filter(
        Order.store_id == store_id,
        Order.status == 'pending',
        or_(
            Order.payment_status == 'paid',
            Order.payment_method == 'cod'
        )
    ).scalar() or 0

    orders_completed = db.query(func.count(Order.id)).filter(
        Order.store_id == store_id,
        Order.status == 'completed',
        func.date(Order.created_at) == date.today()
    ).scalar() or 0

    pending_requests = db.query(func.count(func.distinct(DonationRequest.id))).join(
        DonationRequestItem, DonationRequestItem.request_id == DonationRequest.id
    ).join(
        DonationOffer, DonationOffer.id == DonationRequestItem.offer_id
    ).filter(
        DonationOffer.store_id == store_id,
        DonationRequest.status == 'PENDING'
    ).scalar() or 0

    return {
        "totalLots": int(total_lots),
        "totalInventoryQty": int(total_inventory_qty),
        "nearExpiryProducts": int(near_expiry),
        "lowStockProducts": int(low_stock),
        "ordersToday": int(orders_today),
        "ordersPending": int(orders_pending),
        "ordersCompleted": int(orders_completed),
        "pendingRequests": int(pending_requests),
    }


# ========== Donation Offers Business Logic ==========
def list_near_expiry_lots(db: Session, store_id: int) -> dict:
    today = date.today()
    rows = db.query(
        InventoryLot.id,
        InventoryLot.lot_code,
        InventoryLot.qty_on_hand,
        InventoryLot.expiry_date,
        Product.id.label('product_id'),
        Product.name.label('product_name'),
        Product.base_price,
    ).join(
        Product, Product.id == InventoryLot.product_id
    ).filter(
        InventoryLot.store_id == store_id,
        InventoryLot.qty_on_hand > 0,
        InventoryLot.expiry_date >= today + timedelta(days=3),
        InventoryLot.expiry_date <= today + timedelta(days=30)  # Mở rộng lô gần hết hạn
    ).order_by(
        InventoryLot.expiry_date.asc()
    ).limit(200).all()

    items = []
    for row in rows:
        days_left = (row.expiry_date - today).days
        items.append({
            "id": row.id,
            "productId": row.product_id,
            "productName": row.product_name,
            "lotCode": row.lot_code,
            "quantity": int(row.qty_on_hand),
            "expiryDate": row.expiry_date.strftime("%Y-%m-%d"),
            "daysLeft": days_left,
            "basePrice": float(row.base_price or 0),
        })

    return {"items": items}


def create_donation_offer(db: Session, user_id: int, lot_id: int, offered_qty: int) -> dict:
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    # Validate lot exists and belongs to staff's store
    lot = db.query(
        InventoryLot.id,
        InventoryLot.qty_on_hand,
        InventoryLot.lot_code,
        Product.name.label('product_name'),
        Product.id.label('product_id')
    ).join(
        Product, Product.id == InventoryLot.product_id
    ).filter(
        InventoryLot.id == lot_id,
        InventoryLot.store_id == store_id
    ).first()

    if not lot:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lô hàng không tồn tại hoặc không thuộc cửa hàng"
        )

    if lot.qty_on_hand < offered_qty:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Số lượng donate vượt quá tồn kho"
        )

    if offered_qty <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Số lượng phải lớn hơn 0"
        )

    # Create donation offer
    new_offer = DonationOffer(
        store_id=store_id,
        lot_id=lot_id,
        offered_qty=offered_qty,
        status='open',
        created_by=user_id
    )
    db.add(new_offer)
    db.commit()
    db.refresh(new_offer)

    return {
        "success": True,
        "message": "Tạo donation offer thành công",
        "offerId": new_offer.id,
        "productName": lot.product_name,
        "lotCode": lot.lot_code,
        "offeredQty": offered_qty,
    }


def create_bulk_donation_offers(db: Session, user_id: int, items: list[dict]) -> dict:
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    if not items:
        raise HTTPException(status_code=400, detail="Danh sách sản phẩm không được trống")

    created_offers = []
    errors = []

    for idx, item in enumerate(items):
        lot_id = item.lot_id if not isinstance(item, dict) else item.get("lot_id")
        offered_qty = item.offered_qty if not isinstance(item, dict) else item.get("offered_qty")

        try:
            if not lot_id or not offered_qty:
                errors.append({"row": idx + 1, "message": "Thiếu lot_id hoặc offered_qty"})
                continue

            # Validate lot exists and belongs to staff's store
            lot = db.query(
                InventoryLot.id,
                InventoryLot.qty_on_hand,
                InventoryLot.lot_code,
                Product.name.label('product_name')
            ).join(
                Product, Product.id == InventoryLot.product_id
            ).filter(
                InventoryLot.id == lot_id,
                InventoryLot.store_id == store_id
            ).first()

            if not lot:
                errors.append({"row": idx + 1, "message": f"Lô hàng {lot_id} không tồn tại hoặc không thuộc cửa hàng"})
                continue

            if lot.qty_on_hand < offered_qty:
                errors.append({"row": idx + 1, "message": f"Lô {lot.lot_code}: số lượng donate ({offered_qty}) vượt quá tồn kho ({lot.qty_on_hand})"})
                continue

            if offered_qty <= 0:
                errors.append({"row": idx + 1, "message": f"Lô {lot.lot_code}: số lượng phải lớn hơn 0"})
                continue

            # Create donation offer
            new_offer = DonationOffer(
                store_id=store_id,
                lot_id=lot_id,
                offered_qty=offered_qty,
                status='open',
                created_by=user_id
            )
            db.add(new_offer)
            db.flush()  # Get ID without committing yet

            created_offers.append({
                "offerId": new_offer.id,
                "productName": lot.product_name,
                "lotCode": lot.lot_code,
                "offeredQty": offered_qty,
            })

        except Exception as e:
            errors.append({"row": idx + 1, "message": str(e)})

    if errors:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "message": "Có lỗi khi tạo donation offers",
                "errors": errors
            }
        )

    db.commit()

    return {
        "success": True,
        "message": f"Đã tạo {len(created_offers)} đề nghị quyên góp thành công",
        "created": len(created_offers),
        "offers": created_offers,
    }


def list_staff_donation_offers(db: Session, user_id: int, status_filter: str = "all") -> dict:
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    query = db.query(
        DonationOffer.id,
        DonationOffer.lot_id,
        DonationOffer.offered_qty,
        DonationOffer.status,
        DonationOffer.created_at,
        InventoryLot.lot_code,
        InventoryLot.expiry_date,
        Product.name.label('product_name'),
        func.count(DonationRequestItem.id).label('request_count')
    ).join(
        InventoryLot, InventoryLot.id == DonationOffer.lot_id
    ).join(
        Product, Product.id == InventoryLot.product_id
    ).outerjoin(
        DonationRequestItem, DonationRequestItem.offer_id == DonationOffer.id
    ).outerjoin(
        DonationRequest, DonationRequest.id == DonationRequestItem.request_id
    ).filter(
        DonationOffer.store_id == store_id
    )

    if status_filter != "all":
        query = query.filter(DonationOffer.status == status_filter)

    rows = query.group_by(
        DonationOffer.id, DonationOffer.lot_id, DonationOffer.offered_qty, DonationOffer.status, DonationOffer.created_at,
        InventoryLot.lot_code, InventoryLot.expiry_date, Product.name
    ).order_by(
        DonationOffer.created_at.desc()
    ).limit(200).all()

    items = []
    for row in rows:
        # Calculate remaining quantity from DonationRequestItem (join with DonationRequest)
        total_requested = db.query(func.coalesce(func.sum(DonationRequestItem.quantity), 0)).join(
            DonationRequest, DonationRequest.id == DonationRequestItem.request_id
        ).filter(
            DonationRequestItem.offer_id == row.id,
            DonationRequest.status.in_(['PENDING', 'APPROVED', 'RECEIVED', 'COMPLETED'])
        ).scalar() or 0
        remaining = row.offered_qty - int(total_requested)

        items.append({
            "id": row.id,
            "lotId": row.lot_id,
            "productName": row.product_name,
            "lotCode": row.lot_code,
            "offeredQty": row.offered_qty,
            "remainingQty": max(0, remaining),
            "expiryDate": row.expiry_date.isoformat() if row.expiry_date else None,
            "status": row.status,
            "createdAt": row.created_at.strftime("%d/%m/%Y %H:%M"),
            "requestCount": row.request_count,
        })

    return {"items": items}


def update_donation_offer_status(db: Session, user_id: int, offer_id: int, new_status: str) -> dict:
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    valid_statuses = ['open', 'approved', 'rejected', 'closed']
    if new_status not in valid_statuses:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Trạng thái không hợp lệ"
        )

    offer = db.query(DonationOffer).filter(
        DonationOffer.id == offer_id,
        DonationOffer.store_id == store_id
    ).first()

    if not offer:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Donation offer không tồn tại"
        )

    offer.status = new_status
    db.commit()

    status_text = {
        'approved': 'duyệt',
        'rejected': 'từ chối',
        'closed': 'đóng'
    }.get(new_status, new_status)

    return {
        "success": True,
        "message": f"Đã {status_text} donation offer",
        "status": new_status
    }


def update_donation_offer(db: Session, user_id: int, offer_id: int, offered_qty: int) -> dict:
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    offer = db.query(DonationOffer).filter(
        DonationOffer.id == offer_id,
        DonationOffer.store_id == store_id
    ).first()

    if not offer:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Donation offer không tồn tại"
        )

    if offer.status not in ['open']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Chỉ có thể chỉnh sửa đề nghị đang ở trạng thái Open"
        )

    if offered_qty <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Số lượng phải lớn hơn 0"
        )

    lot = db.query(InventoryLot).filter(InventoryLot.id == offer.lot_id).first()
    if not lot:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lô hàng không tồn tại"
        )

    if offered_qty > lot.qty_on_hand:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Số lượng vượt quá tồn kho (còn lại: {lot.qty_on_hand})"
        )

    offer.offered_qty = offered_qty
    db.commit()

    return {
        "success": True,
        "message": "Đã cập nhật số lượng đề nghị quyên góp",
        "offered_qty": offered_qty
    }


def delete_donation_offer(db: Session, user_id: int, offer_id: int) -> dict:
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    offer = db.query(DonationOffer).filter(
        DonationOffer.id == offer_id,
        DonationOffer.store_id == store_id
    ).first()

    if not offer:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Donation offer không tồn tại"
        )

    if offer.status not in ['open']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Chỉ có thể xóa đề nghị đang ở trạng thái Open"
        )

    lot = db.query(InventoryLot).filter(InventoryLot.id == offer.lot_id).with_for_update().first()
    if lot:
        pending_requests = db.query(DonationRequestItem).join(
            DonationRequest, DonationRequest.id == DonationRequestItem.request_id
        ).filter(
            DonationRequestItem.offer_id == offer_id,
            DonationRequest.status == 'PENDING'
        ).all()

        for req_item in pending_requests:
            lot.qty_reserved = max(0, (lot.qty_reserved or 0) - req_item.quantity)
            req_item.status = 'REJECTED'

    db.delete(offer)
    db.commit()

    return {
        "success": True,
        "message": "Đã xóa đề nghị quyên góp"
    }


# ========== Donation Requests Business Logic (New Architecture) ==========
def list_staff_donation_requests(db: Session, user_id: int, status_filter: str = "all") -> dict:
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    # Query donation requests that have items from offers in this store
    query = db.query(DonationRequest).join(
        DonationRequestItem, DonationRequestItem.request_id == DonationRequest.id
    ).join(
        DonationOffer, DonationOffer.id == DonationRequestItem.offer_id
    ).filter(
        DonationOffer.store_id == store_id
    ).distinct()

    if status_filter != "all":
        query = query.filter(DonationRequest.status == status_filter.upper())

    rows = query.order_by(DonationRequest.created_at.desc()).limit(200).all()

    items = []
    for req in rows:
        # Get charity info
        user = db.query(User).filter(User.id == req.charity_id).first()
        charity_org = db.query(CharityOrganization).filter(
            CharityOrganization.user_id == req.charity_id
        ).first()

        # Count items for this request
        item_count = db.query(func.count(DonationRequestItem.id)).filter(
            DonationRequestItem.request_id == req.id
        ).scalar() or 0

        org_display_name = charity_org.org_name if charity_org and charity_org.org_name else (user.full_name if user else None)

        items.append({
            "id": req.id,
            "charity_id": req.charity_id,
            "organization": org_display_name,
            "charityName": user.full_name if user else None,
            "charityOrgName": charity_org.org_name if charity_org else None,
            "charityPhone": user.phone if user else None,
            "charityAddress": charity_org.address if charity_org else None,
            "status": req.status,
            "total_items": item_count,
            "receivedAt": req.received_at.strftime("%d/%m/%Y %H:%M") if req.received_at else None,
            "createdAt": req.created_at.strftime("%d/%m/%Y %H:%M"),
        })

    return {"items": items}


def get_staff_donation_request_detail(db: Session, user_id: int, request_id: int) -> dict:
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    # Get request and verify it belongs to this store
    request = db.query(DonationRequest).join(
        DonationRequestItem, DonationRequestItem.request_id == DonationRequest.id
    ).join(
        DonationOffer, DonationOffer.id == DonationRequestItem.offer_id
    ).filter(
        DonationRequest.id == request_id,
        DonationOffer.store_id == store_id
    ).first()

    if not request:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Donation request không tồn tại"
        )

    # Get charity info
    user = db.query(User).filter(User.id == request.charity_id).first()
    charity_org = db.query(CharityOrganization).filter(
        CharityOrganization.user_id == request.charity_id
    ).first()

    # Get items with product details
    items_query = db.query(DonationRequestItem).filter(
        DonationRequestItem.request_id == request_id
    ).all()

    items = []
    for item in items_query:
        offer = db.query(DonationOffer).filter(DonationOffer.id == item.offer_id).first()
        product_name = None
        lot_code = None
        expiry_date = None
        store_name = None

        if offer:
            lot = db.query(InventoryLot).filter(InventoryLot.id == offer.lot_id).first()
            if lot:
                product = db.query(Product).filter(Product.id == lot.product_id).first()
                product_name = product.name if product else None
                lot_code = lot.lot_code
                expiry_date = lot.expiry_date.strftime("%d/%m/%Y") if lot.expiry_date else None
            store = db.query(Store).filter(Store.id == offer.store_id).first()
            store_name = store.name if store else None

        items.append({
            "id": item.id,
            "offer_id": item.offer_id,
            "product_name": product_name,
            "lot_code": lot_code,
            "quantity": item.quantity,
            "status": item.status,
            "expiry_date": expiry_date,
            "store_name": store_name,
        })

    org_display_name = charity_org.org_name if charity_org and charity_org.org_name else (user.full_name if user else None)

    return {
        "id": request.id,
        "charity_id": request.charity_id,
        "charity_name": user.full_name if user else None,
        "charity_org_name": charity_org.org_name if charity_org else None,
        "charity_phone": user.phone if user else None,
        "charity_address": charity_org.address if charity_org else None,
        "status": request.status,
        "total_items": len(items),
        "created_at": request.created_at.strftime("%d/%m/%Y %H:%M"),
        "received_at": request.received_at.strftime("%d/%m/%Y %H:%M") if request.received_at else None,
        "items": items,
    }


def update_donation_request_status(db: Session, user_id: int, request_id: int, new_status: str) -> dict:
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    valid_statuses = ['APPROVED', 'REJECTED']
    if new_status not in valid_statuses:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Trạng thái không hợp lệ. Chỉ chấp nhận APPROVED hoặc REJECTED"
        )

    # Get request and verify it belongs to this store
    request = db.query(DonationRequest).join(
        DonationRequestItem, DonationRequestItem.request_id == DonationRequest.id
    ).join(
        DonationOffer, DonationOffer.id == DonationRequestItem.offer_id
    ).filter(
        DonationRequest.id == request_id,
        DonationOffer.store_id == store_id
    ).first()

    if not request:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Donation request không tồn tại"
        )

    if request.status != 'PENDING':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Chỉ có thể xử lý request đang ở trạng thái PENDING"
        )

    if new_status == 'APPROVED':
        # Check stock availability for all items
        items = db.query(DonationRequestItem).filter(
            DonationRequestItem.request_id == request_id
        ).all()

        for item in items:
            offer = db.query(DonationOffer).filter(DonationOffer.id == item.offer_id).first()
            if not offer:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Không tìm thấy offer cho item {item.id}"
                )

            # Calculate total requested (pending + approved) for this offer
            total_requested = db.query(func.sum(DonationRequestItem.quantity)).join(
                DonationRequest, DonationRequest.id == DonationRequestItem.request_id
            ).filter(
                DonationRequestItem.offer_id == item.offer_id,
                DonationRequestItem.id != item.id,
                DonationRequest.status.in_(['PENDING', 'APPROVED', 'RECEIVED', 'COMPLETED'])
            ).scalar() or 0

            available_qty = int(offer.offered_qty or 0) - int(total_requested)
            if item.quantity > available_qty:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Sản phẩm {item.id}: Không đủ số lượng. Còn lại: {available_qty}"
                )

        # Approve request and all items
        db.query(DonationRequest).filter(DonationRequest.id == request_id).update(
            {DonationRequest.status: 'APPROVED'},
            synchronize_session=False
        )
        db.query(DonationRequestItem).filter(
            DonationRequestItem.request_id == request_id
        ).update(
            {DonationRequestItem.status: 'APPROVED'},
            synchronize_session=False
        )

        # Update InventoryLot - Deduct stock immediately as it's being packed/approved
        for item in items:
            offer = db.query(DonationOffer).filter(DonationOffer.id == item.offer_id).first()
            if offer:
                lot = db.query(InventoryLot).filter(InventoryLot.id == offer.lot_id).with_for_update().first()
                if lot:
                    lot.qty_on_hand = max(0, (lot.qty_on_hand or 0) - item.quantity)
                    lot.qty_reserved = max(0, (lot.qty_reserved or 0) - item.quantity)
        
        db.commit()

        # Create delivery for the approved donation request
        try:
            delivery_result = _create_delivery_for_donation_request(db, request_id, store_id)
            delivery_info = f" | Giao hàng: {delivery_result.get('delivery_code', 'N/A')}"
        except HTTPException:
            delivery_info = " | (Khong tao duoc delivery)"

        return {
            "success": True,
            "message": f"Da duyet yeu cau quyen gop{delivery_info}",
            "status": "APPROVED"
        }
    else:
        # Reject request and all items
        items = db.query(DonationRequestItem).filter(DonationRequestItem.request_id == request_id).all()
        for item in items:
            offer = db.query(DonationOffer).filter(DonationOffer.id == item.offer_id).first()
            if offer:
                lot = db.query(InventoryLot).filter(InventoryLot.id == offer.lot_id).with_for_update().first()
                if lot:
                    lot.qty_reserved = max(0, (lot.qty_reserved or 0) - item.quantity)

        db.query(DonationRequest).filter(DonationRequest.id == request_id).update(
            {DonationRequest.status: 'REJECTED'},
            synchronize_session=False
        )
        db.query(DonationRequestItem).filter(
            DonationRequestItem.request_id == request_id
        ).update(
            {DonationRequestItem.status: 'REJECTED'},
            synchronize_session=False
        )
        db.commit()

        return {
            "success": True,
            "message": "Đã từ chối yêu cầu quyên góp và hoàn trả số lượng giữ chỗ",
            "status": "REJECTED"
        }


# ========== File Upload & Import Business Logic ==========
def list_inventory_lots(db: Session, store_id: int, status_filter: str = "all") -> dict:
    from app.services import discount_policy_service
    from datetime import date, timedelta
    
    rows = db.query(
        InventoryLot.id,
        InventoryLot.lot_code,
        InventoryLot.qty_on_hand,
        InventoryLot.qty_reserved,
        InventoryLot.qty_disposed,
        InventoryLot.qty_imported,
        InventoryLot.expiry_date,
        InventoryLot.manufacturing_date,
        InventoryLot.status,
        Product.id.label('product_id'),
        Product.name.label('product_name'),
        Product.base_price,
        Product.supermarket_id
    ).join(
        Product, Product.id == InventoryLot.product_id
    ).filter(
        InventoryLot.store_id == store_id
    ).order_by(
        InventoryLot.expiry_date.asc(), InventoryLot.id.desc()
    ).all()

    items = []
    for row in rows:
        item = _dict_row(row)
        current_status = _normalize_status_value(item.get("status"), item["expiry_date"])
        if status_filter != "all" and status_filter.strip().lower() != current_status.lower():
            continue
        
        # Calculate discount and prices from discount policy
        base_price = float(item["base_price"] or 0)
        expiry_date = item["expiry_date"]
        days_left = (expiry_date - date.today()).days
        
        # Get discount from policy (3-level priority: product > category > supermarket)
        discount_result = discount_policy_service.calculate_discount(
            db,
            base_price,
            expiry_date.strftime("%Y-%m-%d"),
            item["supermarket_id"],
            item["product_id"]
        )
        discount_percent = discount_result.get("discountPercent", 0)
        sale_price = discount_result.get("finalPrice", base_price)
        
        items.append(
            {
                "id": item["id"],
                "lotCode": item["lot_code"],
                "productName": item["product_name"],
                "quantity": int(item["qty_on_hand"] or 0),
                "reserved": int(item["qty_reserved"] or 0),
                "disposed": int(item["qty_disposed"] or 0),
                "imported": max(int(item["qty_imported"] or 0), 
                                int(item["qty_on_hand"] or 0) + int(item["qty_reserved"] or 0) + int(item["qty_disposed"] or 0)),
                "available": max(0, int(item["qty_on_hand"] or 0) - int(item["qty_reserved"] or 0)),
                "manufacturingDate": item["manufacturing_date"].strftime("%Y-%m-%d") if item["manufacturing_date"] else None,
                "expiryDate": expiry_date.strftime("%Y-%m-%d"),
                "status": current_status,
                "basePrice": base_price,
                "salePrice": sale_price,
                "discount": discount_percent,
                "daysLeft": days_left,
            }
        )

    return {"items": items}


def create_inventory_lot(db: Session, store_id: int, supermarket_id: int, lot_code: str, product_name: str,
                        quantity: int, expiry_date: date, manual_status: object, action_note: str,
                        manufacturing_date: date = None) -> dict:
    lot_code = lot_code.strip()
    product_name = product_name.strip()

    if not lot_code or not product_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Du lieu khong hop le")

    try:
        expiry_date = _parse_date_input(expiry_date)
        if manufacturing_date:
            manufacturing_date = _parse_date_input(manufacturing_date)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

    action = _upsert_inventory_lot(
        db,
        store_id=store_id,
        supermarket_id=supermarket_id,
        lot_code=lot_code,
        product_name=product_name,
        quantity=quantity,
        expiry_date=expiry_date,
        manual_status=manual_status,
        manufacturing_date=manufacturing_date,
    )
    db.commit()

    return {"success": True, "action": action, "actionNote": action_note}


def update_inventory_lot(db: Session, lot_id: int, store_id: int, supermarket_id: int, lot_code: str,
                        product_name: str, quantity: int, expiry_date: date, manual_status: object,
                        manufacturing_date: date = None) -> dict:
    lot_code = lot_code.strip()
    product_name = product_name.strip()

    if not lot_code or not product_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Du lieu khong hop le")

    try:
        expiry_date = _parse_date_input(expiry_date)
        if manufacturing_date:
            manufacturing_date = _parse_date_input(manufacturing_date)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

    existing_lot = db.query(InventoryLot).filter(
        InventoryLot.id == lot_id,
        InventoryLot.store_id == store_id
    ).first()
    if not existing_lot:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lô hàng không tồn tại")

    # Kiểm tra nếu trùng mã lô nhưng khác ID (tránh đổi mã lô thành một mã đã có ở record khác)
    duplicate_code = db.query(InventoryLot).filter(
        InventoryLot.lot_code == lot_code,
        InventoryLot.store_id == store_id,
        InventoryLot.id != lot_id
    ).first()
    
    if duplicate_code:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail=f"Mã lô '{lot_code}' đã được sử dụng cho một lô hàng khác trong cửa hàng."
        )

    product_id = _resolve_or_create_product(db, supermarket_id, product_name)
    next_status = _normalize_status_value(manual_status, expiry_date)

    db.query(InventoryLot).filter(
        InventoryLot.id == lot_id,
        InventoryLot.store_id == store_id
    ).update(
        {
            InventoryLot.lot_code: lot_code,
            InventoryLot.product_id: product_id,
            InventoryLot.qty_on_hand: quantity,
            InventoryLot.expiry_date: expiry_date,
            InventoryLot.status: next_status,
            InventoryLot.manufacturing_date: manufacturing_date,
        },
        synchronize_session=False
    )
    db.commit()

    return {"success": True}


def delete_inventory_lot(db: Session, lot_id: int, store_id: int) -> dict:
    lot = db.query(InventoryLot).filter(
        InventoryLot.id == lot_id,
        InventoryLot.store_id == store_id
    ).first()
    
    if not lot:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, 
            detail="Lô hàng không tồn tại"
        )
    
    # Check if lot has reserved quantity
    if lot.qty_reserved > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Không thể xóa lô hàng đang có {lot.qty_reserved} đơn vị được đặt trước. Vui lòng hủy các đơn hàng liên quan trước."
        )
    
    # KIỂM TRA RÀNG BUỘC QUYÊN GÓP: Chặn xóa nếu có Donation Offer
    has_offers = db.query(DonationOffer.id).filter(DonationOffer.lot_id == lot_id).first()
    if has_offers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Không thể xóa lô hàng đang có đề nghị quyên góp. Vui lòng xóa đề nghị quyên góp trước."
        )
    
    db.delete(lot)
    db.commit()

    return {"success": True}


def dispose_inventory_lot(db: Session, user_id: int, lot_id: int, quantity: int, reason: str = "Hết hạn") -> dict:
    """Xác nhận tiêu hủy hàng hết hạn hoặc hư hỏng."""
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    # 1. Tìm lô hàng và khóa để cập nhật
    lot = db.query(InventoryLot).filter(
        InventoryLot.id == lot_id,
        InventoryLot.store_id == store_id
    ).with_for_update().first()

    if not lot:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lô hàng không tồn tại")

    # 2. Kiểm tra số lượng khả dụng để hủy
    available_to_dispose = (lot.qty_on_hand or 0) - (lot.qty_reserved or 0)
    if quantity > available_to_dispose:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Số lượng hủy ({quantity}) vượt quá số lượng khả dụng ({available_to_dispose})"
        )
    
    if quantity <= 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Số lượng hủy phải lớn hơn 0")

    # 3. Cập nhật số lượng
    old_on_hand = lot.qty_on_hand
    old_disposed = lot.qty_disposed or 0
    
    lot.qty_on_hand -= quantity
    lot.qty_disposed = old_disposed + quantity

    # 4. TỰ ĐỘNG CẬP NHẬT DONATION OFFER (Nếu có)
    offer = db.query(DonationOffer).filter(
        DonationOffer.lot_id == lot_id,
        DonationOffer.status == 'open'
    ).first()
    
    if offer:
        # Nếu số lượng sau khi hủy nhỏ hơn số lượng đang rao quyên góp
        if offer.offered_qty > lot.qty_on_hand:
            offer.offered_qty = lot.qty_on_hand
            
        # Nếu kho đã hết sạch hàng khả dụng cho quyên góp -> Đóng Offer
        if offer.offered_qty <= 0:
            offer.status = 'closed'
    
    # 5. Nếu hết sạch hàng thì cập nhật status sang disposed (tùy chọn)
    if lot.qty_on_hand == 0 and lot.qty_reserved == 0:
        lot.status = "disposed"

    db.flush()

    # 4. Ghi nhật ký Audit
    from app.services.audit_service import log_action
    from app.core.audit_actions import DISPOSE_LOT, ENTITY_INVENTORY_LOT
    
    log_action(
        db, 
        user_id=user_id, 
        store_id=store_id,
        action=DISPOSE_LOT,
        entity_type=ENTITY_INVENTORY_LOT,
        entity_id=lot_id,
        old_value={"qty_on_hand": old_on_hand, "qty_disposed": old_disposed},
        new_value={"qty_on_hand": lot.qty_on_hand, "qty_disposed": lot.qty_disposed, "reason": reason}
    )

    db.commit()

    return {
        "success": True, 
        "message": f"Đã tiêu hủy {quantity} sản phẩm. Lý do: {reason}",
        "newQtyOnHand": lot.qty_on_hand,
        "totalDisposed": lot.qty_disposed
    }


# ========== File Upload & Import Business Logic ==========
async def import_inventory_lots_from_excel(db: Session, store_id: int, supermarket_id: int,
                                          file: UploadFile) -> dict:
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File trong")

    rows = _read_import_rows(file, file_bytes)
    if not rows:
        return {
            "success": True,
            "created": 0,
            "updated": 0,
            "failed": 0,
            "errors": [],
            "productsCreated": 0,
            "productsUpdated": 0,
            "lotsCreated": 0,
            "lotsUpdated": 0,
        }

    products_created = 0
    products_updated = 0
    lots_created = 0
    lots_updated = 0
    errors: list[dict[str, object]] = []

    try:
        for idx, row in enumerate(rows, start=2):
            lot_code_raw = _pick_field(row, "lot_code", "lotcode", "ma_lo", "ma_lo_hang")
            product_name_raw = _pick_field(row, "product_name", "product", "name", "ten_san_pham", "san_pham")
            quantity_raw = _pick_field(row, "quantity", "qty", "so_luong")
            expiry_raw = _pick_field(row, "expiry_date", "expiry", "ngay_het_han", "han_su_dung")
            status_raw = _pick_field(row, "status", "trang_thai")
            sku_raw = _pick_field(row, "sku", "ma_sku")
            base_price_raw = _pick_field(row, "base_price", "price", "don_gia", "gia")
            category_raw = _pick_field(row, "category", "category_name", "danh_muc")

            try:
                lot_code = str(lot_code_raw or "").strip()
                product_name = str(product_name_raw or "").strip()
                if not lot_code or not product_name:
                    raise ValueError("Thieu ma lo hoac ten san pham")

                quantity = int(quantity_raw)
                if quantity < 0:
                    raise ValueError("So luong phai >= 0")

                expiry_date = _parse_date_input(expiry_raw)
                product_id, product_action = _upsert_product_from_import(
                    db,
                    supermarket_id=supermarket_id,
                    product_name_raw=product_name,
                    sku_raw=sku_raw,
                    base_price_raw=base_price_raw,
                    category_name_raw=category_raw,
                )
                if product_action == "created":
                    products_created += 1
                else:
                    products_updated += 1

                lot_action = _upsert_inventory_lot(
                    db,
                    store_id=store_id,
                    supermarket_id=supermarket_id,
                    lot_code=lot_code,
                    product_name=product_name,
                    quantity=quantity,
                    expiry_date=expiry_date,
                    manual_status=status_raw,
                    product_id=product_id,
                )

                if lot_action == "created":
                    lots_created += 1
                else:
                    lots_updated += 1
            except Exception as exc:  # noqa: BLE001
                errors.append({"row": idx, "message": str(exc)})
        
        # Commit only if all rows processed (even with some row errors, commit successful ones)
        db.commit()
    except Exception as exc:  # noqa: BLE001
        # Rollback entire transaction on any critical error
        db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Import failed: {str(exc)}")

    return {
        "success": True,
        "created": lots_created,
        "updated": lots_updated,
        "failed": len(errors),
        "errors": errors,
        "productsCreated": products_created,
        "productsUpdated": products_updated,
        "lotsCreated": lots_created,
        "lotsUpdated": lots_updated,
    }


async def import_products_from_excel(db: Session, store_id: int, supermarket_id: int,
                                    file: UploadFile) -> dict:
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File trong")

    rows = _read_import_rows(file, file_bytes)
    if not rows:
        return {
            "success": True,
            "productsCreated": 0,
            "productsUpdated": 0,
            "lotsCreated": 0,
            "lotsUpdated": 0,
            "failed": 0,
            "errors": [],
        }

    products_created = 0
    products_updated = 0
    lots_created = 0
    lots_updated = 0
    errors: list[dict[str, object]] = []

    try:
        for idx, row in enumerate(rows, start=2):
            product_name_raw = _pick_field(row, "product_name", "product", "name", "ten_san_pham", "san_pham")
            sku_raw = _pick_field(row, "sku", "ma_sku")
            base_price_raw = _pick_field(row, "base_price", "price", "don_gia", "gia")
            category_raw = _pick_field(row, "category", "category_name", "danh_muc")

            lot_code_raw = _pick_field(row, "lot_code", "lotcode", "ma_lo", "ma_lo_hang")
            quantity_raw = _pick_field(row, "quantity", "qty", "so_luong")
            expiry_raw = _pick_field(row, "expiry_date", "expiry", "ngay_het_han", "han_su_dung")
            status_raw = _pick_field(row, "status", "trang_thai")

            try:
                product_id, product_action = _upsert_product_from_import(
                    db,
                    supermarket_id=supermarket_id,
                    product_name_raw=product_name_raw,
                    sku_raw=sku_raw,
                    base_price_raw=base_price_raw,
                    category_name_raw=category_raw,
                )
                product_name = str(product_name_raw or "").strip()

                if product_action == "created":
                    products_created += 1
                else:
                    products_updated += 1

                has_lot_data = any(
                    value not in (None, "")
                    for value in (lot_code_raw, quantity_raw, expiry_raw)
                )
                if has_lot_data:
                    lot_code = str(lot_code_raw or "").strip()
                    if not lot_code:
                        raise ValueError("Thieu ma lo de tao ton kho")

                    quantity = int(quantity_raw)
                    if quantity < 0:
                        raise ValueError("So luong phai >= 0")

                    expiry_date = _parse_date_input(expiry_raw)

                    lot_action = _upsert_inventory_lot(
                        db,
                        store_id=store_id,
                        supermarket_id=supermarket_id,
                        lot_code=lot_code,
                        product_name=product_name,
                        quantity=quantity,
                        expiry_date=expiry_date,
                        manual_status=status_raw,
                        product_id=product_id,
                    )
                    if lot_action == "created":
                        lots_created += 1
                    else:
                        lots_updated += 1
            except Exception as exc:  # noqa: BLE001
                errors.append({"row": idx, "message": str(exc)})
        
        # Commit only if all rows processed (even with some row errors, commit successful ones)
        db.commit()
    except Exception as exc:  # noqa: BLE001
        # Rollback entire transaction on any critical error
        db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Import failed: {str(exc)}")

    return {
        "success": True,
        "productsCreated": products_created,
        "productsUpdated": products_updated,
        "lotsCreated": lots_created,
        "lotsUpdated": lots_updated,
        "failed": len(errors),
        "errors": errors,
    }


async def upload_product_image(db: Session, user_id: int, file: UploadFile) -> dict:
    _get_staff_scope(db, user_id)

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Chi chap nhan file hinh anh"
        )

    file_ext = os.path.splitext(file.filename or "")[1] or ".jpg"
    if file_ext.lower() not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Chi chap nhan dinh dang jpg, png, gif, webp"
        )

    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads", "products")
    os.makedirs(uploads_dir, exist_ok=True)

    unique_name = f"{uuid4()}{file_ext}"
    file_path = os.path.join(uploads_dir, unique_name)

    contents = await file.read()
    max_size = 5 * 1024 * 1024
    if len(contents) > max_size:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File qua lon (toi da 5MB)"
        )

    with open(file_path, "wb") as f:
        f.write(contents)

    image_url = f"/uploads/products/{unique_name}"
    return {"url": image_url, "image_url": image_url}
