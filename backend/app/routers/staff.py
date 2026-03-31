from __future__ import annotations

import csv
from datetime import date, datetime
from io import BytesIO, StringIO
import re
import unicodedata

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.database import get_db

router = APIRouter(prefix="/staff", tags=["staff"])


def _dict_row(row) -> dict:
    return dict(row._mapping)


def _status_label(expiry_date: date) -> str:
    today = date.today()
    if expiry_date < today:
        return "Het Han"
    if (expiry_date - today).days <= 7:
        return "Sap Het Han"
    return "Moi"


def _parse_date_input(value) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            raise ValueError("Ngay het han trong")
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(stripped, fmt).date()
            except ValueError:
                continue
    raise ValueError("Ngay het han khong hop le")


def _normalize_header(header: object) -> str:
    if header is None:
        return ""
    value = unicodedata.normalize("NFKD", str(header).strip().lower())
    value = "".join(char for char in value if not unicodedata.combining(char))
    value = value.replace(" ", "_")
    value = value.replace("-", "_")
    return value


def _normalize_status_value(raw_status: object, expiry_date: date) -> str:
    if raw_status is None:
        return _status_label(expiry_date)

    normalized = unicodedata.normalize("NFKD", str(raw_status).strip().lower())
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = normalized.replace("_", " ")

    status_map = {
        "moi": "Moi",
        "sap het han": "Sap Het Han",
        "het han": "Het Han",
    }
    return status_map.get(normalized, _status_label(expiry_date))


def _parse_non_negative_float(value: object, field_name: str) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None

    normalized = str(value).strip().replace(" ", "")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(",", "")
    elif "," in normalized and "." not in normalized:
        normalized = normalized.replace(",", ".")
    try:
        result = float(normalized)
    except ValueError as exc:
        raise ValueError(f"{field_name} khong hop le") from exc

    if result < 0:
        raise ValueError(f"{field_name} phai >= 0")
    return result


def _build_product_sku(name: str) -> str:
    base = re.sub(r"[^a-z0-9]+", "", name.strip().lower())
    if not base:
        base = "product"
    return f"sku_{base[:20]}"


def _generate_unique_sku(db: Session, supermarket_id: int, product_name: str) -> str:
    base = _build_product_sku(product_name)
    candidate = base
    index = 1

    while True:
        exists = db.execute(
            text(
                """
                SELECT id
                FROM products
                WHERE supermarket_id = :supermarket_id
                  AND sku = :sku
                LIMIT 1
                """
            ),
            {"supermarket_id": supermarket_id, "sku": candidate},
        ).first()

        if not exists:
            return candidate

        index += 1
        candidate = f"{base}_{index}"


def _get_staff_scope(db: Session, user_id: int) -> dict[str, int]:
    user = db.execute(
        text(
            """
            SELECT id, role, store_id, supermarket_id
            FROM users
            WHERE id = :id
            LIMIT 1
            """
        ),
        {"id": user_id},
    ).first()

    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    role = (user.role or "").lower()
    if role != "store_staff":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Tai khoan khong co quyen staff.",
        )

    if not user.store_id or not user.supermarket_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tai khoan staff chua duoc gan store.",
        )

    return {"store_id": int(user.store_id), "supermarket_id": int(user.supermarket_id)}


def _resolve_or_create_category(db: Session, category_name: object) -> int | None:
    name = str(category_name or "").strip()
    if not name:
        return None

    existing = db.execute(
        text(
            """
            SELECT id
            FROM categories
            WHERE LOWER(name) = LOWER(:name)
            LIMIT 1
            """
        ),
        {"name": name},
    ).first()

    if existing:
        return int(existing.id)

    result = db.execute(
        text("INSERT INTO categories (name) VALUES (:name)"),
        {"name": name},
    )
    return int(result.lastrowid)


def _ensure_unique_sku_for_product(
    db: Session,
    supermarket_id: int,
    sku: str,
    current_product_id: int | None = None,
) -> None:
    conflict = db.execute(
        text(
            """
            SELECT id
            FROM products
            WHERE supermarket_id = :supermarket_id
              AND sku = :sku
              AND (:current_product_id IS NULL OR id != :current_product_id)
            LIMIT 1
            """
        ),
        {
            "supermarket_id": supermarket_id,
            "sku": sku,
            "current_product_id": current_product_id,
        },
    ).first()

    if conflict:
        raise ValueError(f"SKU '{sku}' da ton tai")


def _upsert_product_from_import(
    db: Session,
    *,
    supermarket_id: int,
    product_name_raw: object,
    sku_raw: object = None,
    base_price_raw: object = None,
    category_name_raw: object = None,
) -> tuple[int, str]:
    product_name = str(product_name_raw or "").strip()
    if not product_name:
        raise ValueError("Ten san pham khong duoc trong")

    sku = str(sku_raw or "").strip()
    if not sku:
        sku = None

    base_price = _parse_non_negative_float(base_price_raw, "Don gia")
    category_id = _resolve_or_create_category(db, category_name_raw)

    product = None
    if sku:
        product = db.execute(
            text(
                """
                SELECT id, sku, base_price, category_id
                FROM products
                WHERE supermarket_id = :supermarket_id
                  AND sku = :sku
                LIMIT 1
                """
            ),
            {"supermarket_id": supermarket_id, "sku": sku},
        ).first()

    if not product:
        product = db.execute(
            text(
                """
                SELECT id, sku, base_price, category_id
                FROM products
                WHERE supermarket_id = :supermarket_id
                  AND LOWER(name) = LOWER(:name)
                LIMIT 1
                """
            ),
            {"supermarket_id": supermarket_id, "name": product_name},
        ).first()

    if product:
        current_id = int(product.id)
        next_sku = str(product.sku)
        if sku and sku != next_sku:
            _ensure_unique_sku_for_product(db, supermarket_id, sku, current_product_id=current_id)
            next_sku = sku

        next_base_price = float(base_price) if base_price is not None else float(product.base_price or 0)
        next_category_id = category_id if category_id is not None else product.category_id

        db.execute(
            text(
                """
                UPDATE products
                SET sku = :sku,
                    name = :name,
                    base_price = :base_price,
                    category_id = :category_id
                WHERE id = :id
                """
            ),
            {
                "sku": next_sku,
                "name": product_name,
                "base_price": next_base_price,
                "category_id": next_category_id,
                "id": current_id,
            },
        )
        return current_id, "updated"

    next_sku = sku or _generate_unique_sku(db, supermarket_id, product_name)
    _ensure_unique_sku_for_product(db, supermarket_id, next_sku)
    result = db.execute(
        text(
            """
            INSERT INTO products (supermarket_id, category_id, sku, name, base_price, image_url)
            VALUES (:supermarket_id, :category_id, :sku, :name, :base_price, NULL)
            """
        ),
        {
            "supermarket_id": supermarket_id,
            "category_id": category_id,
            "sku": next_sku,
            "name": product_name,
            "base_price": float(base_price) if base_price is not None else 0,
        },
    )
    return int(result.lastrowid), "created"


def _resolve_or_create_product(db: Session, supermarket_id: int, product_name: str) -> int:
    name = product_name.strip()
    product = db.execute(
        text(
            """
            SELECT id
            FROM products
            WHERE supermarket_id = :supermarket_id
              AND LOWER(name) = LOWER(:name)
            LIMIT 1
            """
        ),
        {"supermarket_id": supermarket_id, "name": name},
    ).first()

    if product:
        return int(product.id)

    sku = _generate_unique_sku(db, supermarket_id, name)
    result = db.execute(
        text(
            """
            INSERT INTO products (supermarket_id, category_id, sku, name, base_price, image_url)
            VALUES (:supermarket_id, NULL, :sku, :name, 0, NULL)
            """
        ),
        {"supermarket_id": supermarket_id, "sku": sku, "name": name},
    )
    return int(result.lastrowid)


def _upsert_inventory_lot(
    db: Session,
    *,
    store_id: int,
    supermarket_id: int,
    lot_code: str,
    product_name: str,
    quantity: int,
    expiry_date: date,
    manual_status: object = None,
    product_id: int | None = None,
) -> str:
    lot = db.execute(
        text(
            """
            SELECT id
            FROM inventory_lots
            WHERE store_id = :store_id
              AND lot_code = :lot_code
            LIMIT 1
            """
        ),
        {"store_id": store_id, "lot_code": lot_code},
    ).first()

    if product_id is None:
        product_id = _resolve_or_create_product(db, supermarket_id, product_name)
    lot_status = _normalize_status_value(manual_status, expiry_date)

    if lot:
        db.execute(
            text(
                """
                UPDATE inventory_lots
                SET product_id = :product_id,
                    expiry_date = :expiry_date,
                    qty_on_hand = :qty_on_hand,
                    status = :status
                WHERE id = :id
                """
            ),
            {
                "product_id": product_id,
                "expiry_date": expiry_date,
                "qty_on_hand": quantity,
                "status": lot_status,
                "id": int(lot.id),
            },
        )
        return "updated"

    db.execute(
        text(
            """
            INSERT INTO inventory_lots
                (store_id, product_id, lot_code, expiry_date, qty_on_hand, qty_reserved, status)
            VALUES
                (:store_id, :product_id, :lot_code, :expiry_date, :qty_on_hand, 0, :status)
            """
        ),
        {
            "store_id": store_id,
            "product_id": product_id,
            "lot_code": lot_code,
            "expiry_date": expiry_date,
            "qty_on_hand": quantity,
            "status": lot_status,
        },
    )
    return "created"


def _extract_rows_from_xlsx(file_bytes: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_header(h) for h in rows[0]]
    data_rows: list[dict[str, object]] = []
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item: dict[str, object] = {}
        for idx, value in enumerate(row):
            if idx < len(headers) and headers[idx]:
                item[headers[idx]] = value
        data_rows.append(item)
    return data_rows


def _extract_rows_from_csv(file_bytes: bytes) -> list[dict[str, object]]:
    text_stream = StringIO(file_bytes.decode("utf-8-sig"))
    reader = csv.DictReader(text_stream)
    rows = []
    for row in reader:
        normalized = {_normalize_header(k): v for k, v in row.items()}
        if any(str(v or "").strip() for v in normalized.values()):
            rows.append(normalized)
    return rows


def _read_import_rows(upload: UploadFile, file_bytes: bytes) -> list[dict[str, object]]:
    filename = (upload.filename or "").lower()
    if filename.endswith(".xlsx"):
        return _extract_rows_from_xlsx(file_bytes)
    if filename.endswith(".csv"):
        return _extract_rows_from_csv(file_bytes)

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Chi ho tro file .xlsx hoac .csv",
    )


def _pick_field(row: dict[str, object], *keys: str) -> object:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


@router.get("/profile")
def get_staff_profile(
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    user = db.execute(
        text(
            """
            SELECT u.id, u.username, u.email, u.full_name, u.phone, u.role,
                   u.store_id, u.supermarket_id,
                   s.name AS store_name, s.location AS store_address
            FROM users u
            LEFT JOIN stores s ON s.id = u.store_id
            WHERE u.id = :user_id
            LIMIT 1
            """
        ),
        {"user_id": user_id},
    ).first()

    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    return {
        "email": user.email,
        "fullName": user.full_name,
        "phone": user.phone,
        "role": user.role,
        "storeName": user.store_name or "Chưa gán cửa hàng",
        "storeAddress": user.store_address or "",
    }


@router.put("/profile")
def update_staff_profile(
    payload: dict,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    full_name = (payload.get("fullName") or "").strip()
    email = (payload.get("email") or "").strip().lower()
    phone = (payload.get("phone") or "").strip()

    if not full_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Họ tên không được trống")
    if not email:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email không được trống")

    existing = db.execute(
        text(
            """
            SELECT id FROM users
            WHERE email = :email AND id != :user_id
            LIMIT 1
            """
        ),
        {"email": email, "user_id": user_id},
    ).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email đã được sử dụng")

    db.execute(
        text(
            """
            UPDATE users
            SET full_name = :full_name,
                email = :email,
                phone = :phone
            WHERE id = :user_id
            """
        ),
        {"full_name": full_name, "email": email, "phone": phone, "user_id": user_id},
    )
    db.commit()
    return {"success": True}


@router.get("/orders")
def list_staff_orders(
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    rows = db.execute(
        text(
            """
            SELECT o.id, o.status, o.total_amount, o.payment_status, o.created_at,
                   u.full_name AS customer_name
            FROM orders o
            JOIN users u ON u.id = o.customer_id
            WHERE o.store_id = :store_id
            ORDER BY o.created_at DESC
            LIMIT 100
            """
        ),
        {"store_id": store_id},
    ).all()

    items = []
    for row in rows:
        items.append({
            "id": f"DH-{row.id}",
            "orderId": row.id,
            "customer": row.customer_name,
            "status": row.status,
            "amount": f"{float(row.total_amount or 0):,.0f} VNĐ" if row.total_amount else "0 VNĐ",
            "paymentStatus": row.payment_status,
            "createdAt": row.created_at.strftime("%d/%m/%Y %H:%M"),
        })

    return {"items": items}


@router.put("/orders/{order_id}/status")
def update_staff_order_status(
    order_id: int,
    payload: dict,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    new_status = (payload.get("status") or "").strip().lower()

    valid_statuses = ["pending", "preparing", "ready", "completed", "cancelled"]
    if new_status not in valid_statuses:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Trạng thái không hợp lệ")

    result = db.execute(
        text(
            """
            UPDATE orders
            SET status = :status
            WHERE id = :order_id AND store_id = :store_id
            """
        ),
        {"status": new_status, "order_id": order_id, "store_id": scope["store_id"]},
    )
    db.commit()

    if (result.rowcount or 0) == 0:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Đơn hàng không tồn tại")

    return {"success": True, "status": new_status}


@router.get("/notifications")
def list_staff_notifications(
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    rows = db.execute(
        text(
            """
            SELECT id, type, content, is_read, created_at
            FROM notifications
            WHERE user_id = :user_id
            ORDER BY created_at DESC
            LIMIT 50
            """
        ),
        {"user_id": user_id},
    ).all()

    items = []
    for row in rows:
        items.append({
            "id": row.id,
            "type": row.type or "info",
            "content": row.content or "",
            "isRead": bool(row.is_read),
            "createdAt": row.created_at.strftime("%d/%m/%Y %H:%M"),
        })

    return {"items": items}


@router.put("/notifications/{notification_id}/read")
def mark_notification_as_read(
    notification_id: int,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    result = db.execute(
        text(
            """
            UPDATE notifications
            SET is_read = 1
            WHERE id = :notification_id AND user_id = :user_id
            """
        ),
        {"notification_id": notification_id, "user_id": user_id},
    )
    db.commit()

    return {"success": True}


@router.get("/category-stats")
def staff_category_stats(
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    rows = db.execute(
        text(
            """
            SELECT COALESCE(c.name, 'Khác') AS category_name,
                   COUNT(il.id) AS lot_count
            FROM inventory_lots il
            JOIN products p ON p.id = il.product_id
            LEFT JOIN categories c ON c.id = p.category_id
            WHERE il.store_id = :store_id
              AND il.expiry_date >= CURDATE()
              AND il.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 7 DAY)
            GROUP BY c.id, c.name
            ORDER BY lot_count DESC
            LIMIT 10
            """
        ),
        {"store_id": store_id},
    ).all()

    total = sum(row.lot_count for row in rows) or 1
    items = []
    for row in rows:
        items.append({
            "name": row.category_name,
            "percent": int((row.lot_count / total) * 100),
        })

    return {"items": items}


@router.get("/categories")
def list_categories(
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    supermarket_id = scope["supermarket_id"]

    rows = db.execute(
        text(
            """
            SELECT c.id, c.name, COUNT(p.id) AS product_count
            FROM categories c
            LEFT JOIN products p ON p.category_id = c.id AND p.supermarket_id = :supermarket_id
            GROUP BY c.id, c.name
            ORDER BY c.name ASC
            """
        ),
        {"supermarket_id": supermarket_id},
    ).all()

    items = [
        {"id": row.id, "name": row.name, "productCount": row.product_count}
        for row in rows
    ]

    return {"items": items}


@router.post("/categories")
def create_category(
    payload: dict,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên danh mục không được trống")

    existing = db.execute(
        text("SELECT id FROM categories WHERE LOWER(name) = LOWER(:name) LIMIT 1"),
        {"name": name},
    ).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Danh mục đã tồn tại")

    db.execute(
        text("INSERT INTO categories (name) VALUES (:name)"),
        {"name": name},
    )
    db.commit()

    return {"message": "Tạo danh mục thành công", "name": name}


@router.put("/categories/{category_id}")
def update_category(
    category_id: int,
    payload: dict,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên danh mục không được trống")

    existing = db.execute(
        text("SELECT id FROM categories WHERE LOWER(name) = LOWER(:name) AND id != :id LIMIT 1"),
        {"name": name, "id": category_id},
    ).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Danh mục đã tồn tại")

    db.execute(
        text("UPDATE categories SET name = :name WHERE id = :id"),
        {"name": name, "id": category_id},
    )
    db.commit()

    return {"message": "Cập nhật danh mục thành công"}


@router.delete("/categories/{category_id}")
def delete_category(
    category_id: int,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    has_products = db.execute(
        text("SELECT COUNT(*) FROM products WHERE category_id = :category_id"),
        {"category_id": category_id},
    ).scalar() or 0

    if has_products > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Không thể xóa: có {has_products} sản phẩm đang sử dụng danh mục này"
        )

    db.execute(text("DELETE FROM categories WHERE id = :id"), {"id": category_id})
    db.commit()

    return {"message": "Xóa danh mục thành công"}


@router.get("/products")
def list_products(
    user_id: int = Query(..., ge=1),
    category_filter: int = Query(default=None),
    search: str = Query(default=None),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    supermarket_id = scope["supermarket_id"]

    query = """
        SELECT p.id, p.sku, p.name, p.base_price, p.image_url,
               c.name AS category_name, c.id AS category_id,
               COALESCE(SUM(il.qty_on_hand), 0) AS total_stock
        FROM products p
        LEFT JOIN categories c ON c.id = p.category_id
        LEFT JOIN inventory_lots il ON il.product_id = p.id
        WHERE p.supermarket_id = :supermarket_id
    """
    params = {"supermarket_id": supermarket_id}

    if category_filter is not None:
        query += " AND p.category_id = :category_id"
        params["category_id"] = category_filter

    if search:
        query += " AND (p.name LIKE :search OR p.sku LIKE :search)"
        params["search"] = f"%{search}%"

    query += " GROUP BY p.id, p.sku, p.name, p.base_price, p.image_url, c.name, c.id ORDER BY p.name ASC"

    rows = db.execute(text(query), params).all()

    items = [
        {
            "id": row.id,
            "sku": row.sku,
            "name": row.name,
            "basePrice": float(row.base_price),
            "imageUrl": row.image_url,
            "categoryName": row.category_name or "Chưa phân loại",
            "categoryId": row.category_id,
            "totalStock": int(row.total_stock),
        }
        for row in rows
    ]

    return {"items": items}


@router.post("/products")
def create_product(
    payload: dict,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    supermarket_id = scope["supermarket_id"]

    name = (payload.get("name") or "").strip()
    sku = (payload.get("sku") or "").strip()
    base_price = payload.get("basePrice")
    category_id = payload.get("categoryId")
    image_url = (payload.get("imageUrl") or "").strip() or None

    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên sản phẩm không được trống")
    if not sku:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SKU không được trống")
    if base_price is None or float(base_price) < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Giá không hợp lệ")

    existing = db.execute(
        text("SELECT id FROM products WHERE supermarket_id = :sm AND sku = :sku LIMIT 1"),
        {"sm": supermarket_id, "sku": sku},
    ).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SKU đã tồn tại")

    db.execute(
        text(
            "INSERT INTO products (supermarket_id, category_id, sku, name, base_price, image_url) "
            "VALUES (:sm, :cat, :sku, :name, :price, :img)"
        ),
        {
            "sm": supermarket_id,
            "cat": category_id if category_id else None,
            "sku": sku,
            "name": name,
            "price": float(base_price),
            "img": image_url,
        },
    )
    db.commit()

    return {"message": "Tạo sản phẩm thành công"}


@router.put("/products/{product_id}")
def update_product(
    product_id: int,
    payload: dict,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    supermarket_id = scope["supermarket_id"]

    name = (payload.get("name") or "").strip()
    base_price = payload.get("basePrice")
    category_id = payload.get("categoryId")
    image_url = (payload.get("imageUrl") or "").strip() or None

    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên sản phẩm không được trống")
    if base_price is None or float(base_price) < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Giá không hợp lệ")

    product = db.execute(
        text("SELECT id FROM products WHERE id = :id AND supermarket_id = :sm LIMIT 1"),
        {"id": product_id, "sm": supermarket_id},
    ).first()
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sản phẩm không tìm thấy")

    db.execute(
        text(
            "UPDATE products SET name = :name, base_price = :price, "
            "category_id = :cat, image_url = :img WHERE id = :id"
        ),
        {
            "name": name,
            "price": float(base_price),
            "cat": category_id if category_id else None,
            "img": image_url,
            "id": product_id,
        },
    )
    db.commit()

    return {"message": "Cập nhật sản phẩm thành công"}


@router.delete("/products/{product_id}")
def delete_product(
    product_id: int,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    supermarket_id = scope["supermarket_id"]

    product = db.execute(
        text("SELECT id FROM products WHERE id = :id AND supermarket_id = :sm LIMIT 1"),
        {"id": product_id, "sm": supermarket_id},
    ).first()
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sản phẩm không tìm thấy")

    has_lots = db.execute(
        text("SELECT COUNT(*) FROM inventory_lots WHERE product_id = :product_id"),
        {"product_id": product_id},
    ).scalar() or 0

    if has_lots > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Không thể xóa: có {has_lots} lô hàng liên quan đến sản phẩm này"
        )

    db.execute(text("DELETE FROM products WHERE id = :id"), {"id": product_id})
    db.commit()

    return {"message": "Xóa sản phẩm thành công"}


@router.get("/products/categories")
def list_product_categories(
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    supermarket_id = scope["supermarket_id"]

    rows = db.execute(
        text(
            """
            SELECT DISTINCT c.id, c.name
            FROM categories c
            LEFT JOIN products p ON p.category_id = c.id AND p.supermarket_id = :sm
            WHERE c.id IN (SELECT category_id FROM products WHERE supermarket_id = :sm)
               OR c.id NOT IN (SELECT DISTINCT category_id FROM products WHERE supermarket_id = :sm AND category_id IS NOT NULL)
            ORDER BY c.name ASC
            """
        ),
        {"sm": supermarket_id},
    ).all()

    items = [{"id": row.id, "name": row.name} for row in rows]
    return {"items": items}


@router.get("/dashboard-summary")
def staff_dashboard_summary(
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    total_lots = db.execute(
        text("SELECT COUNT(*) FROM inventory_lots WHERE store_id = :store_id"),
        {"store_id": store_id},
    ).scalar() or 0

    near_expiry = db.execute(
        text(
            """
            SELECT COUNT(*)
            FROM inventory_lots
            WHERE store_id = :store_id
              AND expiry_date >= CURDATE()
              AND expiry_date <= DATE_ADD(CURDATE(), INTERVAL 7 DAY)
              AND qty_on_hand > 0
            """
        ),
        {"store_id": store_id},
    ).scalar() or 0

    orders_today = db.execute(
        text(
            """
            SELECT COUNT(*)
            FROM orders
            WHERE store_id = :store_id
              AND DATE(created_at) = CURDATE()
            """
        ),
        {"store_id": store_id},
    ).scalar() or 0

    pending_requests = db.execute(
        text(
            """
            SELECT COUNT(*)
            FROM donation_requests dr
            JOIN donation_offers dof ON dof.id = dr.offer_id
            WHERE dof.store_id = :store_id
              AND LOWER(dr.status) = 'pending'
            """
        ),
        {"store_id": store_id},
    ).scalar() or 0

    return {
        "totalLots": int(total_lots),
        "nearExpiryProducts": int(near_expiry),
        "ordersToday": int(orders_today),
        "pendingRequests": int(pending_requests),
    }


@router.get("/inventory-lots")
def list_inventory_lots(
    user_id: int = Query(..., ge=1),
    status_filter: str = Query(default="all"),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)
    store_id = scope["store_id"]

    rows = db.execute(
        text(
            """
            SELECT
                il.id,
                il.lot_code,
                il.qty_on_hand,
                il.expiry_date,
                il.status,
                p.name AS product_name
            FROM inventory_lots il
            JOIN products p ON p.id = il.product_id
            WHERE il.store_id = :store_id
            ORDER BY il.expiry_date ASC, il.id DESC
            """
        ),
        {"store_id": store_id},
    ).all()

    items = []
    for row in rows:
        item = _dict_row(row)
        current_status = _normalize_status_value(item.get("status"), item["expiry_date"])
        if status_filter != "all" and status_filter.strip().lower() != current_status.lower():
            continue
        items.append(
            {
                "id": item["id"],
                "lotCode": item["lot_code"],
                "productName": item["product_name"],
                "quantity": int(item["qty_on_hand"] or 0),
                "expiryDate": item["expiry_date"].strftime("%Y-%m-%d"),
                "status": current_status,
            }
        )

    return {"items": items}


@router.post("/inventory-lots")
def create_inventory_lot(
    payload: dict,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)

    lot_code = (payload.get("lotCode") or "").strip()
    product_name = (payload.get("productName") or "").strip()
    quantity_raw = payload.get("quantity")
    expiry_date_raw = payload.get("expiryDate")
    manual_status = payload.get("status")
    action_note = (payload.get("actionNote") or "").strip()

    if not lot_code or not product_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Du lieu khong hop le")

    try:
        quantity = int(quantity_raw)
        if quantity < 0:
            raise ValueError
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="So luong khong hop le") from exc

    try:
        expiry_date = _parse_date_input(expiry_date_raw)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

    action = _upsert_inventory_lot(
        db,
        store_id=scope["store_id"],
        supermarket_id=scope["supermarket_id"],
        lot_code=lot_code,
        product_name=product_name,
        quantity=quantity,
        expiry_date=expiry_date,
        manual_status=manual_status,
    )
    db.commit()

    return {"success": True, "action": action, "actionNote": action_note}


@router.put("/inventory-lots/{lot_id}")
def update_inventory_lot(
    lot_id: int,
    payload: dict,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)

    exists = db.execute(
        text(
            """
            SELECT id
            FROM inventory_lots
            WHERE id = :lot_id
              AND store_id = :store_id
            LIMIT 1
            """
        ),
        {"lot_id": lot_id, "store_id": scope["store_id"]},
    ).first()
    if not exists:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lo hang khong ton tai")

    lot_code = (payload.get("lotCode") or "").strip()
    product_name = (payload.get("productName") or "").strip()
    quantity_raw = payload.get("quantity")
    expiry_date_raw = payload.get("expiryDate")
    manual_status = payload.get("status")

    if not lot_code or not product_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Du lieu khong hop le")

    try:
        quantity = int(quantity_raw)
        if quantity < 0:
            raise ValueError
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="So luong khong hop le") from exc

    try:
        expiry_date = _parse_date_input(expiry_date_raw)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

    product_id = _resolve_or_create_product(db, scope["supermarket_id"], product_name)
    next_status = _normalize_status_value(manual_status, expiry_date)

    db.execute(
        text(
            """
            UPDATE inventory_lots
            SET lot_code = :lot_code,
                product_id = :product_id,
                qty_on_hand = :qty_on_hand,
                expiry_date = :expiry_date,
                status = :status
            WHERE id = :lot_id
              AND store_id = :store_id
            """
        ),
        {
            "lot_code": lot_code,
            "product_id": product_id,
            "qty_on_hand": quantity,
            "expiry_date": expiry_date,
            "status": next_status,
            "lot_id": lot_id,
            "store_id": scope["store_id"],
        },
    )
    db.commit()

    return {"success": True}


@router.delete("/inventory-lots/{lot_id}")
def delete_inventory_lot(
    lot_id: int,
    user_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)

    result = db.execute(
        text(
            """
            DELETE FROM inventory_lots
            WHERE id = :lot_id
              AND store_id = :store_id
            """
        ),
        {"lot_id": lot_id, "store_id": scope["store_id"]},
    )
    db.commit()

    if (result.rowcount or 0) == 0:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lo hang khong ton tai")

    return {"success": True}


@router.post("/inventory-lots/import-excel")
async def import_inventory_lots_from_excel(
    user_id: int = Query(..., ge=1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File trong")

    rows = _read_import_rows(file, file_bytes)
    if not rows:
        return {
            "success": True,
            "created": 0,
            "updated": 0,
            "failed": 0,
            "errors": [],
            "productsCreated": 0,
            "productsUpdated": 0,
            "lotsCreated": 0,
            "lotsUpdated": 0,
        }

    products_created = 0
    products_updated = 0
    lots_created = 0
    lots_updated = 0
    errors: list[dict[str, object]] = []

    for idx, row in enumerate(rows, start=2):
        lot_code_raw = _pick_field(row, "lot_code", "lotcode", "ma_lo", "ma_lo_hang")
        product_name_raw = _pick_field(row, "product_name", "product", "name", "ten_san_pham", "san_pham")
        quantity_raw = _pick_field(row, "quantity", "qty", "so_luong")
        expiry_raw = _pick_field(row, "expiry_date", "expiry", "ngay_het_han", "han_su_dung")
        status_raw = _pick_field(row, "status", "trang_thai")
        sku_raw = _pick_field(row, "sku", "ma_sku")
        base_price_raw = _pick_field(row, "base_price", "price", "don_gia", "gia")
        category_raw = _pick_field(row, "category", "category_name", "danh_muc")

        try:
            lot_code = str(lot_code_raw or "").strip()
            product_name = str(product_name_raw or "").strip()
            if not lot_code or not product_name:
                raise ValueError("Thieu ma lo hoac ten san pham")

            quantity = int(quantity_raw)
            if quantity < 0:
                raise ValueError("So luong phai >= 0")

            expiry_date = _parse_date_input(expiry_raw)
            product_id, product_action = _upsert_product_from_import(
                db,
                supermarket_id=scope["supermarket_id"],
                product_name_raw=product_name,
                sku_raw=sku_raw,
                base_price_raw=base_price_raw,
                category_name_raw=category_raw,
            )
            if product_action == "created":
                products_created += 1
            else:
                products_updated += 1

            lot_action = _upsert_inventory_lot(
                db,
                store_id=scope["store_id"],
                supermarket_id=scope["supermarket_id"],
                lot_code=lot_code,
                product_name=product_name,
                quantity=quantity,
                expiry_date=expiry_date,
                manual_status=status_raw,
                product_id=product_id,
            )

            if lot_action == "created":
                lots_created += 1
            else:
                lots_updated += 1
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": idx, "message": str(exc)})

    db.commit()

    return {
        "success": True,
        "created": lots_created,
        "updated": lots_updated,
        "failed": len(errors),
        "errors": errors,
        "productsCreated": products_created,
        "productsUpdated": products_updated,
        "lotsCreated": lots_created,
        "lotsUpdated": lots_updated,
    }


@router.post("/products/import-excel")
async def import_products_from_excel(
    user_id: int = Query(..., ge=1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    scope = _get_staff_scope(db, user_id)

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File trong")

    rows = _read_import_rows(file, file_bytes)
    if not rows:
        return {
            "success": True,
            "productsCreated": 0,
            "productsUpdated": 0,
            "lotsCreated": 0,
            "lotsUpdated": 0,
            "failed": 0,
            "errors": [],
        }

    products_created = 0
    products_updated = 0
    lots_created = 0
    lots_updated = 0
    errors: list[dict[str, object]] = []

    for idx, row in enumerate(rows, start=2):
        product_name_raw = _pick_field(row, "product_name", "product", "name", "ten_san_pham", "san_pham")
        sku_raw = _pick_field(row, "sku", "ma_sku")
        base_price_raw = _pick_field(row, "base_price", "price", "don_gia", "gia")
        category_raw = _pick_field(row, "category", "category_name", "danh_muc")

        lot_code_raw = _pick_field(row, "lot_code", "lotcode", "ma_lo", "ma_lo_hang")
        quantity_raw = _pick_field(row, "quantity", "qty", "so_luong")
        expiry_raw = _pick_field(row, "expiry_date", "expiry", "ngay_het_han", "han_su_dung")
        status_raw = _pick_field(row, "status", "trang_thai")

        try:
            product_id, product_action = _upsert_product_from_import(
                db,
                supermarket_id=scope["supermarket_id"],
                product_name_raw=product_name_raw,
                sku_raw=sku_raw,
                base_price_raw=base_price_raw,
                category_name_raw=category_raw,
            )
            product_name = str(product_name_raw or "").strip()

            if product_action == "created":
                products_created += 1
            else:
                products_updated += 1

            has_lot_data = any(
                value not in (None, "")
                for value in (lot_code_raw, quantity_raw, expiry_raw)
            )
            if has_lot_data:
                lot_code = str(lot_code_raw or "").strip()
                if not lot_code:
                    raise ValueError("Thieu ma lo de tao ton kho")

                quantity = int(quantity_raw)
                if quantity < 0:
                    raise ValueError("So luong phai >= 0")

                expiry_date = _parse_date_input(expiry_raw)

                lot_action = _upsert_inventory_lot(
                    db,
                    store_id=scope["store_id"],
                    supermarket_id=scope["supermarket_id"],
                    lot_code=lot_code,
                    product_name=product_name,
                    quantity=quantity,
                    expiry_date=expiry_date,
                    manual_status=status_raw,
                    product_id=product_id,
                )
                if lot_action == "created":
                    lots_created += 1
                else:
                    lots_updated += 1
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": idx, "message": str(exc)})

    db.commit()

    return {
        "success": True,
        "productsCreated": products_created,
        "productsUpdated": products_updated,
        "lotsCreated": lots_created,
        "lotsUpdated": lots_updated,
        "failed": len(errors),
        "errors": errors,
    }
